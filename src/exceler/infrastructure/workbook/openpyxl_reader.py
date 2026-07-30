"""OpenPyxl adapter for factual workbook inspection (Phase 2A).

Open options (documented):
- data_only=False — preserve formula expressions; never evaluate.
- keep_vba=False — do not load VBA for execution; VBA presence is detected via ZIP.
- keep_links=True — retain external link metadata if present (never followed/downloaded).
- read_only=False — full load so tables, dimensions, merges and styles are observable.

Pathological dimensions:
When declared_area (max_row * max_column) exceeds the remaining max_cells_scanned budget,
the adapter inspects only openpyxl-materialized cells (worksheet._cells), encapsulated here.
That path never walks the full declared rectangle. See docs/limitations.md.
"""

from __future__ import annotations

import logging
import time
import uuid
import zipfile
from collections.abc import Iterable
from datetime import UTC, date, datetime
from datetime import time as time_cls
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from exceler.application.workbook.ports import WorkbookSource
from exceler.domain.workbook.enums import (
    CellValueKind,
    InspectionCompletionStatus,
    InspectionTruncationCode,
    InspectionWarningCode,
    WorkbookFormat,
    WorksheetVisibility,
)
from exceler.domain.workbook.errors import (
    EncryptedWorkbookError,
    InvalidWorkbookError,
    UnsupportedWorkbookFormatError,
    WorkbookLimitExceededError,
)
from exceler.domain.workbook.models import (
    INSPECTOR_VERSION,
    CellComment,
    CellInspection,
    CellValue,
    ColumnDimensionInspection,
    DefinedNameInspection,
    ExternalLinkInspection,
    FileIdentity,
    HyperlinkInspection,
    InspectionTruncation,
    InspectionWarning,
    MergedRangeInspection,
    RelevantCellStyle,
    RowDimensionInspection,
    StructuredTableColumnInspection,
    StructuredTableInspection,
    WorkbookInspection,
    WorkbookInspectionOptions,
    WorksheetInspection,
)

logger = logging.getLogger("exceler.workbook.inspection")


def _has_vba_project(payload: bytes) -> bool:
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            return any(name.lower().endswith("vbaproject.bin") for name in archive.namelist())
    except zipfile.BadZipFile:
        return False


def _map_visibility(state: str | None) -> WorksheetVisibility:
    if state == "hidden":
        return WorksheetVisibility.HIDDEN
    if state == "veryHidden":
        return WorksheetVisibility.VERY_HIDDEN
    return WorksheetVisibility.VISIBLE


def _map_value(raw: Any, *, library_data_type: str | None) -> CellValue:
    if raw is None:
        return CellValue(kind=CellValueKind.NULL)
    if isinstance(raw, bool):
        return CellValue(kind=CellValueKind.BOOLEAN, boolean=raw)
    if isinstance(raw, int) and not isinstance(raw, bool):
        return CellValue(kind=CellValueKind.INTEGER, integer=raw)
    if isinstance(raw, float):
        return CellValue(kind=CellValueKind.DECIMAL, decimal=format(Decimal(str(raw)), "f"))
    if isinstance(raw, Decimal):
        return CellValue(kind=CellValueKind.DECIMAL, decimal=format(raw, "f"))
    if isinstance(raw, datetime):
        return CellValue(kind=CellValueKind.DATETIME, datetime=raw.isoformat())
    if isinstance(raw, date):
        return CellValue(kind=CellValueKind.DATE, date=raw.isoformat())
    if isinstance(raw, time_cls):
        return CellValue(kind=CellValueKind.TIME, time=raw.isoformat())
    if library_data_type == "e" or (
        isinstance(raw, str) and raw.startswith("#") and raw.endswith("!")
    ):
        return CellValue(kind=CellValueKind.ERROR, error=str(raw))
    if isinstance(raw, str):
        return CellValue(kind=CellValueKind.STRING, text=raw)
    return CellValue(kind=CellValueKind.STRING, text=str(raw))


def _color_token(color: Any) -> str | None:
    """Stable factual color token when openpyxl exposes rgb/theme/indexed."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and rgb and rgb != "00000000":
        return rgb.upper() if len(rgb) in {6, 8} else rgb
    theme = getattr(color, "theme", None)
    if theme is not None:
        return f"theme:{theme}"
    indexed = getattr(color, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"
    return None


def _border_present(side: Any) -> bool:
    if side is None:
        return False
    style = getattr(side, "style", None)
    return style is not None and style != "none"


def _extract_style(cell: Any) -> RelevantCellStyle | None:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border

    font_name = getattr(font, "name", None) if font is not None else None
    font_size = getattr(font, "size", None) if font is not None else None
    font_bold = bool(font and font.bold)
    font_color = _color_token(getattr(font, "color", None) if font is not None else None)

    fill_color = None
    if fill is not None:
        pattern = getattr(fill, "patternType", None) or getattr(fill, "fill_type", None)
        if pattern and pattern != "none":
            fill_color = _color_token(getattr(fill, "fgColor", None)) or _color_token(
                getattr(fill, "start_color", None)
            )

    horizontal = None
    if alignment is not None and alignment.horizontal:
        horizontal = str(alignment.horizontal)

    border_top = _border_present(getattr(border, "top", None) if border is not None else None)
    border_right = _border_present(getattr(border, "right", None) if border is not None else None)
    border_bottom = _border_present(getattr(border, "bottom", None) if border is not None else None)
    border_left = _border_present(getattr(border, "left", None) if border is not None else None)

    number_format = cell.number_format if cell.number_format not in (None, "General") else None

    has_signal = any(
        [
            font_name is not None,
            font_size is not None,
            font_bold,
            font_color is not None,
            fill_color is not None,
            horizontal is not None,
            border_top,
            border_right,
            border_bottom,
            border_left,
            number_format is not None,
        ]
    )
    if not has_signal:
        return None
    return RelevantCellStyle(
        font_name=font_name,
        font_size=float(font_size) if font_size is not None else None,
        font_bold=font_bold,
        font_color=font_color,
        fill_color=fill_color,
        horizontal_alignment=horizontal,
        border_top=border_top,
        border_right=border_right,
        border_bottom=border_bottom,
        border_left=border_left,
        number_format=number_format,
    )


def _cell_relevant(cell: Any, *, options: WorkbookInspectionOptions) -> bool:
    if cell.value is not None:
        return True
    if options.include_comments and cell.comment is not None:
        return True
    if options.include_hyperlinks and cell.hyperlink is not None:
        return True
    if cell.data_type == "f":
        return True
    if options.include_empty_formatted_cells:
        if cell.number_format not in (None, "General"):
            return True
        if cell.font is not None and cell.font.bold:
            return True
        style = _extract_style(cell)
        if style is not None and (
            style.fill_color is not None
            or style.border_top
            or style.border_right
            or style.border_bottom
            or style.border_left
        ):
            return True
    return False


def _iter_materialized_cells(ws: Any) -> Iterable[Any]:
    """Yield openpyxl-materialized cells only (never the full declared rectangle).

    Encapsulates worksheet._cells — an openpyxl internal structure — so pathological
    dimensions do not force O(max_row * max_column) iteration.
    """
    cells_map = getattr(ws, "_cells", None)
    if not isinstance(cells_map, dict):
        return
    for _coord, cell in sorted(cells_map.items(), key=lambda item: (item[0][0], item[0][1])):
        yield cell


def _build_cell_inspection(cell: Any, *, options: WorkbookInspectionOptions) -> CellInspection:
    formula = None
    raw = cell.value
    if isinstance(raw, str) and raw.startswith("="):
        formula = raw
        value = CellValue(kind=CellValueKind.NULL)
    elif cell.data_type == "f" and raw is not None:
        formula = str(raw)
        value = CellValue(kind=CellValueKind.NULL)
    else:
        value = _map_value(raw, library_data_type=cell.data_type)
    comment = None
    if options.include_comments and cell.comment is not None:
        comment = CellComment(text=cell.comment.text, author=cell.comment.author)
    hyperlink = None
    if options.include_hyperlinks and cell.hyperlink is not None:
        hyperlink = HyperlinkInspection(
            target=cell.hyperlink.target,
            tooltip=getattr(cell.hyperlink, "tooltip", None),
        )
    style = _extract_style(cell)
    return CellInspection(
        coordinate=cell.coordinate,
        row=cell.row,
        column=cell.column,
        value=value,
        library_data_type=cell.data_type,
        number_format=cell.number_format,
        formula=formula,
        comment=comment,
        hyperlink=hyperlink,
        style=style,
    )


class OpenPyxlWorkbookReader:
    def inspect(
        self,
        source: WorkbookSource,
        options: WorkbookInspectionOptions | None = None,
    ) -> WorkbookInspection:
        opts = options or WorkbookInspectionOptions()
        started = time.perf_counter()
        inspected_at = datetime.now(tz=UTC)
        inspection_id = str(uuid.uuid4())

        logger.info(
            "inspection_started",
            extra={
                "event": "inspection_started",
                "file_name": source.name,
                "inspection_id": inspection_id,
            },
        )

        announced_size = source.size_bytes()
        if announced_size > opts.max_file_size_bytes:
            raise WorkbookLimitExceededError(
                f"Workbook exceeds max_file_size_bytes ({opts.max_file_size_bytes})."
            )
        if announced_size == 0:
            raise InvalidWorkbookError("Workbook file is empty.")

        extension = source.suggested_extension.lstrip(".").lower()
        if extension not in {"xlsx", "xlsm"}:
            raise UnsupportedWorkbookFormatError(
                f"Unsupported workbook format: {extension or 'unknown'}"
            )
        workbook_format = WorkbookFormat.XLSM if extension == "xlsm" else WorkbookFormat.XLSX

        with source.open_binary() as handle:
            payload = handle.read()

        # Identity is bound to the exact bytes handed to openpyxl — never a second read.
        content_hash = sha256(payload).hexdigest()
        actual_size = len(payload)
        if actual_size == 0:
            raise InvalidWorkbookError("Workbook file is empty.")
        if actual_size > opts.max_file_size_bytes:
            raise WorkbookLimitExceededError(
                f"Workbook payload exceeds max_file_size_bytes ({opts.max_file_size_bytes})."
            )

        has_vba = _has_vba_project(payload)
        warnings: list[InspectionWarning] = []
        truncations: list[InspectionTruncation] = []
        limitations: list[str] = [
            "Formulas are observed, not evaluated (data_only=False).",
            "VBA projects are detected by ZIP presence only; never executed.",
            "External links are listed when present; never followed or downloaded.",
            "Document core/custom properties are not required for Phase 2A.",
            "FileIdentity.content_hash is sha256 of the inspected payload bytes.",
            "Pathological declared dimensions use materialized-cell fallback (adapter-internal).",
            "Cell styles expose factual presence signals (font/fill/alignment/borders) for Phase 2B.",
        ]

        if announced_size != actual_size:
            warnings.append(
                InspectionWarning(
                    code=InspectionWarningCode.SOURCE_SIZE_CHANGED,
                    message=(
                        "Announced source size differs from inspected payload length; "
                        "FileIdentity uses payload length and payload hash."
                    ),
                )
            )

        if has_vba:
            warnings.append(
                InspectionWarning(
                    code=InspectionWarningCode.VBA_PROJECT_PRESENT,
                    message="Workbook package contains vbaProject.bin (not executed).",
                )
            )

        try:
            wb = load_workbook(
                BytesIO(payload),
                read_only=False,
                data_only=False,
                keep_vba=False,
                keep_links=True,
            )
        except InvalidFileException as exc:
            message = str(exc).lower()
            if "olecf" in message or "encrypted" in message or "password" in message:
                raise EncryptedWorkbookError() from exc
            raise InvalidWorkbookError() from exc
        except zipfile.BadZipFile as exc:
            raise InvalidWorkbookError() from exc
        except KeyError as exc:
            raise InvalidWorkbookError() from exc
        except Exception as exc:  # noqa: BLE001
            raise InvalidWorkbookError() from exc

        logger.debug(
            "workbook_opened",
            extra={
                "event": "workbook_opened",
                "inspection_id": inspection_id,
                "format": workbook_format.value,
                "sheet_count": len(wb.sheetnames),
                "size_bytes": actual_size,
            },
        )

        if len(wb.sheetnames) > opts.max_worksheets:
            wb.close()
            raise WorkbookLimitExceededError(
                f"Workbook exceeds max_worksheets ({opts.max_worksheets})."
            )

        worksheets: list[WorksheetInspection] = []
        cells_observed_total = 0
        cells_scanned_total = 0
        observe_budget = opts.max_cells_observed
        scan_budget = opts.max_cells_scanned

        for index, name in enumerate(wb.sheetnames):
            if observe_budget <= 0 or scan_budget <= 0:
                truncations.append(
                    InspectionTruncation(
                        code=(
                            InspectionTruncationCode.MAX_CELLS_OBSERVED
                            if observe_budget <= 0
                            else InspectionTruncationCode.MAX_CELLS_SCANNED
                        ),
                        message="Remaining worksheets skipped due to cell budget.",
                        location=name,
                    )
                )
                break
            ws = wb[name]
            sheet_inspection, observed, scanned, sheet_truncations = self._inspect_worksheet(
                ws,
                index=index,
                options=opts,
                observe_budget=observe_budget,
                scan_budget=scan_budget,
                warnings=warnings,
            )
            worksheets.append(sheet_inspection)
            cells_observed_total += observed
            cells_scanned_total += scanned
            observe_budget -= observed
            scan_budget -= scanned
            truncations.extend(sheet_truncations)
            logger.debug(
                "worksheet_inspected",
                extra={
                    "event": "worksheet_inspected",
                    "inspection_id": inspection_id,
                    "sheet": name,
                    "cells_observed": observed,
                    "cells_scanned": scanned,
                },
            )

        defined_names = tuple(
            sorted(
                (
                    DefinedNameInspection(
                        name=key,
                        attr_text=getattr(wb.defined_names[key], "attr_text", None),
                        local_sheet_id=getattr(wb.defined_names[key], "localSheetId", None),
                    )
                    for key in wb.defined_names.keys()
                ),
                key=lambda item: (item.local_sheet_id is not None, item.name),
            )
        )

        external_links: list[ExternalLinkInspection] = []
        if opts.include_external_links:
            for link in getattr(wb, "_external_links", None) or []:
                target = getattr(getattr(link, "file_link", None), "Target", None)
                external_links.append(ExternalLinkInspection(target=str(target or link)))
            if external_links:
                warnings.append(
                    InspectionWarning(
                        code=InspectionWarningCode.EXTERNAL_LINK_PRESENT,
                        message="Workbook declares external links (not followed).",
                    )
                )

        wb.close()

        modified_at = None
        modified_iso = source.modified_at_iso()
        if modified_iso:
            modified_at = datetime.fromisoformat(modified_iso)

        # Deduplicate truncation codes while preserving order.
        seen_trunc: set[str] = set()
        unique_truncations: list[InspectionTruncation] = []
        for item in truncations:
            key = f"{item.code.value}:{item.location}"
            if key in seen_trunc:
                continue
            seen_trunc.add(key)
            unique_truncations.append(item)

        completion = (
            InspectionCompletionStatus.PARTIAL
            if unique_truncations
            else InspectionCompletionStatus.COMPLETE
        )
        duration_ms = int((time.perf_counter() - started) * 1000)
        result = WorkbookInspection(
            inspection_id=inspection_id,
            inspector_version=INSPECTOR_VERSION,
            inspected_at=inspected_at,
            duration_ms=duration_ms,
            format=workbook_format,
            file=FileIdentity(
                source_path=source.source_path(),
                file_name=source.name,
                extension=extension,
                size_bytes=actual_size,
                modified_at=modified_at,
                content_hash=content_hash,
            ),
            worksheets=tuple(worksheets),
            defined_names=defined_names,
            external_links=tuple(sorted(external_links, key=lambda item: item.target)),
            has_vba_project=has_vba,
            warnings=tuple(warnings),
            completion_status=completion,
            truncation_reasons=tuple(unique_truncations),
            limitations=tuple(limitations),
            cells_observed=cells_observed_total,
            cells_scanned=cells_scanned_total,
            worksheets_observed=len(worksheets),
        )
        logger.info(
            "inspection_completed",
            extra={
                "event": "inspection_completed",
                "inspection_id": inspection_id,
                "duration_ms": duration_ms,
                "cells_observed": cells_observed_total,
                "cells_scanned": cells_scanned_total,
                "completion_status": completion.value,
                "has_vba_project": has_vba,
            },
        )
        return result

    def _inspect_worksheet(
        self,
        ws: Any,
        *,
        index: int,
        options: WorkbookInspectionOptions,
        observe_budget: int,
        scan_budget: int,
        warnings: list[InspectionWarning],
    ) -> tuple[WorksheetInspection, int, int, list[InspectionTruncation]]:
        truncations: list[InspectionTruncation] = []
        declared = None
        try:
            declared = ws.calculate_dimension()
        except Exception:  # noqa: BLE001
            declared = None

        auto_filter = None
        if ws.auto_filter is not None and ws.auto_filter.ref:
            auto_filter = ws.auto_filter.ref

        merged: list[MergedRangeInspection] = []
        for rng in sorted(str(r) for r in ws.merged_cells.ranges):
            anchor = rng.split(":")[0]
            cell = ws[anchor]
            raw = cell.value
            if isinstance(raw, str) and raw.startswith("="):
                value = CellValue(kind=CellValueKind.NULL)
            else:
                value = _map_value(raw, library_data_type=getattr(cell, "data_type", None))
            merged.append(MergedRangeInspection(reference=rng, anchor=anchor, anchor_value=value))

        row_dims: list[RowDimensionInspection] = []
        for idx, dim in sorted(ws.row_dimensions.items(), key=lambda item: int(item[0])):
            if dim.hidden or dim.height is not None:
                row_dims.append(
                    RowDimensionInspection(
                        index=int(idx),
                        hidden=bool(dim.hidden),
                        height=float(dim.height) if dim.height is not None else None,
                    )
                )

        col_dims: list[ColumnDimensionInspection] = []
        for letter, dim in sorted(ws.column_dimensions.items()):
            if dim.hidden or dim.width is not None:
                col_dims.append(
                    ColumnDimensionInspection(
                        letter=letter,
                        hidden=bool(dim.hidden),
                        width=float(dim.width) if dim.width is not None else None,
                    )
                )

        tables: list[StructuredTableInspection] = []
        for name in sorted(ws.tables.keys()):
            table = ws.tables[name]
            columns: list[StructuredTableColumnInspection] = []
            table_columns = getattr(table, "tableColumns", None)
            if table_columns is not None:
                for col_index, col in enumerate(list(table_columns)):
                    columns.append(
                        StructuredTableColumnInspection(
                            name=getattr(col, "name", f"Column{col_index + 1}"),
                            index=col_index,
                        )
                    )
            auto = None
            if table.autoFilter is not None:
                auto = table.autoFilter.ref
            tables.append(
                StructuredTableInspection(
                    name=name,
                    display_name=table.displayName,
                    reference=table.ref,
                    header_row_count=int(table.headerRowCount or 0),
                    totals_row_count=int(table.totalsRowCount or 0),
                    auto_filter=auto,
                    columns=tuple(columns),
                )
            )

        max_row = int(ws.max_row or 1)
        max_col = int(ws.max_column or 1)
        declared_area = max_row * max_col
        use_materialized = declared_area > scan_budget

        cells: list[CellInspection] = []
        scanned = 0
        observed = 0

        if use_materialized:
            warnings.append(
                InspectionWarning(
                    code=InspectionWarningCode.MATERIALIZED_CELLS_FALLBACK,
                    message=(
                        "Declared dimension area exceeds max_cells_scanned; "
                        "inspecting materialized cells only."
                    ),
                    location=ws.title,
                )
            )
            truncations.append(
                InspectionTruncation(
                    code=InspectionTruncationCode.MAX_CELLS_SCANNED,
                    message="Declared sheet area exceeds scan budget; full rectangle not walked.",
                    location=ws.title,
                )
            )
            for cell in _iter_materialized_cells(ws):
                if scanned >= scan_budget:
                    break
                scanned += 1
                if not _cell_relevant(cell, options=options):
                    continue
                if observed >= observe_budget:
                    truncations.append(
                        InspectionTruncation(
                            code=InspectionTruncationCode.MAX_CELLS_OBSERVED,
                            message="max_cells_observed reached while collecting cells.",
                            location=ws.title,
                        )
                    )
                    break
                cells.append(_build_cell_inspection(cell, options=options))
                observed += 1
        else:
            stop_observe = False
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if scanned >= scan_budget:
                        truncations.append(
                            InspectionTruncation(
                                code=InspectionTruncationCode.MAX_CELLS_SCANNED,
                                message="max_cells_scanned reached during rectangle walk.",
                                location=ws.title,
                            )
                        )
                        stop_observe = True
                        break
                    scanned += 1
                    if not _cell_relevant(cell, options=options):
                        continue
                    if observed >= observe_budget:
                        truncations.append(
                            InspectionTruncation(
                                code=InspectionTruncationCode.MAX_CELLS_OBSERVED,
                                message="max_cells_observed reached while collecting cells.",
                                location=ws.title,
                            )
                        )
                        stop_observe = True
                        break
                    cells.append(_build_cell_inspection(cell, options=options))
                    observed += 1
                if stop_observe:
                    break

        cells.sort(key=lambda item: (item.row, item.column))

        if declared and ":" in declared:
            valued_rows = [
                c.row for c in cells if c.formula is not None or c.value.kind.value != "null"
            ]
            valued_cols = [
                c.column for c in cells if c.formula is not None or c.value.kind.value != "null"
            ]
            last_row = max(valued_rows, default=1)
            last_col = max(valued_cols, default=1)
            if max_row > last_row * 5 + 50 or max_col > last_col * 5 + 10:
                warnings.append(
                    InspectionWarning(
                        code=InspectionWarningCode.DIMENSION_MAY_BE_INFLATED,
                        message=(
                            "Worksheet declared_dimension appears larger than "
                            "observed valued cells."
                        ),
                        location=ws.title,
                    )
                )

        return (
            WorksheetInspection(
                name=ws.title,
                index=index,
                visibility=_map_visibility(ws.sheet_state),
                declared_dimension=declared,
                freeze_panes=ws.freeze_panes,
                auto_filter=auto_filter,
                merged_ranges=tuple(merged),
                row_dimensions=tuple(row_dims),
                column_dimensions=tuple(col_dims),
                tables=tuple(tables),
                cells=tuple(cells),
                cells_observed=observed,
                cells_scanned=scanned,
            ),
            observed,
            scanned,
            truncations,
        )
