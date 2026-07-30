from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from tests.generators.catalog import ALL_SPECS, get_builder
from tests.generators.generate_fixtures import generate_all
from tests.generators.structural_scenarios import (
    build_excel_structured_table,
    build_formulas,
    build_hidden_columns,
    build_hidden_sheet,
    build_simple_rectangular_table,
    build_xlsm_container,
)
from tests.generators.verify_fixtures import logical_snapshot, verify_all
from tests.generators.workbook_factory import (
    EXPECTED_SCHEMA_VERSION,
    ScenarioSpec,
    expected_path,
    fixtures_root,
    manifest_path,
    save_workbook,
    workbook_path,
    write_expected_skeleton,
    write_index,
    write_manifest,
)

pytestmark = pytest.mark.unit

DEFAULT_SEED = 20260730


def _spec(
    scenario_id: str,
    *,
    category: str = "minimal",
    workbook: str | None = None,
    generator: str = "structural_scenarios.build_simple_rectangular_table",
    seed: int | None = DEFAULT_SEED,
    features: list[str] | None = None,
    expected_skeleton: dict[str, Any] | None = None,
) -> ScenarioSpec:
    return ScenarioSpec(
        scenario_id=scenario_id,
        category=category,
        description=f"Synthetic isolated scenario {scenario_id}",
        relative_workbook=workbook or f"workbooks/{category}/{scenario_id}.xlsx",
        generator_name=generator,
        seed=seed,
        features=features or [],
        expected_skeleton=expected_skeleton or {"worksheets": 1},
    )


def _materialize(
    root: Path,
    specs: list[ScenarioSpec],
    *,
    builders: dict[str, Any] | None = None,
) -> None:
    resolve = builders or {}
    for spec in specs:
        builder = resolve.get(spec.generator_name) or get_builder(spec.generator_name)
        assert spec.seed is not None
        wb = builder(spec.seed)
        save_workbook(wb, workbook_path(spec, root=root))
        write_manifest(spec, root=root)
        write_expected_skeleton(spec, root=root)
    write_index(specs, root=root)


def test_catalog_has_minimum_coverage() -> None:
    assert len(ALL_SPECS) >= 10
    scenario_ids = {spec.scenario_id for spec in ALL_SPECS}
    assert "hell_erp_clientes" in scenario_ids
    assert "hell_erp_facturas" in scenario_ids
    assert "simple_rectangular_table" in scenario_ids
    assert "xlsm_container" in scenario_ids


def test_generate_and_verify_roundtrip() -> None:
    generate_all()
    errors = verify_all()
    assert errors == []


def test_manifests_match_workbooks() -> None:
    generate_all()
    for spec in ALL_SPECS:
        man = json.loads(manifest_path(spec).read_text(encoding="utf-8"))
        wb = workbook_path(spec)
        assert wb.exists()
        assert man["workbook"].endswith(wb.name)
        assert man["seed"] == spec.seed
        loaded = load_workbook(wb, read_only=True, data_only=False)
        assert loaded.sheetnames
        loaded.close()


def test_expected_skeletons_have_schema_version() -> None:
    generate_all()
    for spec in ALL_SPECS:
        payload = json.loads(expected_path(spec).read_text(encoding="utf-8"))
        assert payload["schema_version"] == EXPECTED_SCHEMA_VERSION
        assert payload["scenario_id"] == spec.scenario_id
        assert payload["workbook"] == spec.relative_workbook.replace("\\", "/")
        assert isinstance(payload["expectations"], dict)


def test_fixtures_stay_under_fixtures_root() -> None:
    root = fixtures_root().resolve()
    for spec in ALL_SPECS:
        path = workbook_path(spec).resolve()
        assert root in path.parents or path == root
        assert ".." not in Path(spec.relative_workbook).parts


def test_xlsm_container_has_no_vba_project() -> None:
    generate_all()
    path = workbook_path(next(s for s in ALL_SPECS if s.scenario_id == "xlsm_container"))
    with zipfile.ZipFile(path) as archive:
        names = [name.lower() for name in archive.namelist()]
    assert not any(name.endswith("vbaproject.bin") for name in names)


def test_logical_snapshot_captures_hidden_column(tmp_path: Path) -> None:
    tmp = tmp_path / "hidden.xlsx"
    save_workbook(build_hidden_columns(DEFAULT_SEED), tmp)
    snap = logical_snapshot(tmp)
    assert "B" in snap["sheets"][0]["column_dimensions"]["hidden"]


def test_verify_detects_duplicate_scenario_id(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec_a = _spec("dup")
    spec_b = _spec("dup", workbook="workbooks/minimal/dup_b.xlsx")
    _materialize(root, [spec_a])
    errors = verify_all(specs=[spec_a, spec_b], root=root)
    assert any("Duplicate scenario_id" in item for item in errors)


def test_verify_detects_duplicate_workbook_path(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    shared = "workbooks/minimal/shared.xlsx"
    spec_a = _spec("one", workbook=shared)
    spec_b = _spec("two", workbook=shared)
    _materialize(root, [spec_a])
    errors = verify_all(specs=[spec_a, spec_b], root=root)
    assert any("Duplicate workbook path" in item for item in errors)


def test_verify_detects_missing_seed(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("noseed", seed=None)
    seeded = _spec("noseed", seed=DEFAULT_SEED)
    _materialize(root, [seeded])
    errors = verify_all(specs=[spec], root=root)
    assert any("seed is missing" in item for item in errors)


def test_verify_detects_missing_manifest(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("nomanifest")
    _materialize(root, [spec])
    manifest_path(spec, root=root).unlink()
    errors = verify_all(specs=[spec], root=root)
    assert any("missing manifest" in item for item in errors)


def test_verify_detects_missing_expected(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("noexpected")
    _materialize(root, [spec])
    expected_path(spec, root=root).unlink()
    errors = verify_all(specs=[spec], root=root)
    assert any("missing expected" in item for item in errors)


def test_verify_detects_missing_workbook(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("nowb")
    _materialize(root, [spec])
    workbook_path(spec, root=root).unlink()
    errors = verify_all(specs=[spec], root=root)
    assert any("missing workbook" in item for item in errors)


def test_verify_detects_invalid_manifest_json(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("badjson")
    _materialize(root, [spec])
    manifest_path(spec, root=root).write_text("{not-json", encoding="utf-8")
    errors = verify_all(specs=[spec], root=root)
    assert any("invalid manifest JSON" in item for item in errors)


def test_verify_detects_catalog_manifest_mismatch(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("mismatch")
    _materialize(root, [spec])
    payload = json.loads(manifest_path(spec, root=root).read_text(encoding="utf-8"))
    payload["generator"] = "structural_scenarios.build_hidden_sheet"
    manifest_path(spec, root=root).write_text(
        json.dumps(payload, indent=2) + "\n", encoding="utf-8"
    )
    errors = verify_all(specs=[spec], root=root)
    assert any("manifest generator mismatch" in item for item in errors)


def test_verify_detects_path_escape(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("escape", workbook="../outside.xlsx")
    write_manifest(spec, root=root)
    write_expected_skeleton(spec, root=root)
    write_index([spec], root=root)
    target = workbook_path(spec, root=root)
    target.parent.mkdir(parents=True, exist_ok=True)
    save_workbook(build_simple_rectangular_table(DEFAULT_SEED), target)
    errors = verify_all(specs=[spec], root=root)
    assert any("escapes fixtures root" in item for item in errors)


def test_verify_detects_unreadable_workbook(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("broken")
    _materialize(root, [spec])
    workbook_path(spec, root=root).write_bytes(b"not-an-xlsx")
    errors = verify_all(specs=[spec], root=root)
    assert any("not readable" in item for item in errors)


def test_verify_detects_orphan_workbook(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("kept")
    _materialize(root, [spec])
    orphan = root / "workbooks" / "minimal" / "orphan.xlsx"
    save_workbook(build_simple_rectangular_table(DEFAULT_SEED), orphan)
    errors = verify_all(specs=[spec], root=root)
    assert any("Orphan workbook" in item for item in errors)


def test_verify_detects_orphan_manifest(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("kept2")
    _materialize(root, [spec])
    orphan = root / "manifests" / "minimal" / "orphan.json"
    orphan.write_text("{}", encoding="utf-8")
    errors = verify_all(specs=[spec], root=root)
    assert any("Orphan manifest" in item for item in errors)


def test_verify_detects_orphan_expected(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("kept3")
    _materialize(root, [spec])
    orphan = root / "expected" / "minimal" / "orphan.json"
    orphan.write_text("{}", encoding="utf-8")
    errors = verify_all(specs=[spec], root=root)
    assert any("Orphan expected" in item for item in errors)


def test_verify_detects_logical_diff_after_regen(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec(
        "logic",
        generator="structural_scenarios.build_simple_rectangular_table",
    )
    builders = {
        "structural_scenarios.build_simple_rectangular_table": build_simple_rectangular_table
    }
    _materialize(root, [spec], builders=builders)
    wb = load_workbook(workbook_path(spec, root=root))
    wb.active["A1"] = "Changed"
    save_workbook(wb, workbook_path(spec, root=root))
    errors = verify_all(specs=[spec], root=root, builders=builders)
    assert any("logical content differs" in item for item in errors)


def test_verify_detects_lost_hidden_column(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec(
        "hidecol",
        category="structural",
        generator="structural_scenarios.build_hidden_columns",
        features=["hidden_columns"],
    )
    builders = {"structural_scenarios.build_hidden_columns": build_hidden_columns}
    _materialize(root, [spec], builders=builders)
    wb = load_workbook(workbook_path(spec, root=root))
    wb.active.column_dimensions["B"].hidden = False
    save_workbook(wb, workbook_path(spec, root=root))
    errors = verify_all(specs=[spec], root=root, builders=builders)
    assert any("logical content differs" in item for item in errors)


def test_verify_detects_lost_hidden_sheet(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec(
        "hidesheet",
        category="structural",
        generator="structural_scenarios.build_hidden_sheet",
        features=["hidden_sheet"],
    )
    builders = {"structural_scenarios.build_hidden_sheet": build_hidden_sheet}
    _materialize(root, [spec], builders=builders)
    wb = load_workbook(workbook_path(spec, root=root))
    wb["HiddenData"].sheet_state = "visible"
    save_workbook(wb, workbook_path(spec, root=root))
    errors = verify_all(specs=[spec], root=root, builders=builders)
    assert any("logical content differs" in item for item in errors)


def test_verify_detects_formula_change(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec(
        "formula",
        category="types",
        generator="structural_scenarios.build_formulas",
        features=["formulas"],
    )
    builders = {"structural_scenarios.build_formulas": build_formulas}
    _materialize(root, [spec], builders=builders)
    wb = load_workbook(workbook_path(spec, root=root))
    wb.active["C2"] = "=A2+B2"
    save_workbook(wb, workbook_path(spec, root=root))
    errors = verify_all(specs=[spec], root=root, builders=builders)
    assert any("logical content differs" in item for item in errors)


def test_verify_detects_table_ref_change(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec(
        "table",
        category="structural",
        generator="structural_scenarios.build_excel_structured_table",
        features=["excel_table"],
    )
    builders = {"structural_scenarios.build_excel_structured_table": build_excel_structured_table}
    _materialize(root, [spec], builders=builders)
    wb = load_workbook(workbook_path(spec, root=root))
    wb.active.tables["InventoryTable"].ref = "A1:B3"
    save_workbook(wb, workbook_path(spec, root=root))
    errors = verify_all(specs=[spec], root=root, builders=builders)
    assert any("logical content differs" in item for item in errors)


def test_verify_detects_defined_name_change(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec(
        "xlsm",
        category="structural",
        workbook="workbooks/structural/xlsm.xlsx",
        generator="structural_scenarios.build_xlsm_container",
        features=["xlsm", "no_vba_project"],
    )
    builders = {"structural_scenarios.build_xlsm_container": build_xlsm_container}
    _materialize(root, [spec], builders=builders)
    wb = load_workbook(workbook_path(spec, root=root), keep_vba=False)
    del wb.defined_names["FixtureFlag"]
    save_workbook(wb, workbook_path(spec, root=root))
    errors = verify_all(specs=[spec], root=root, builders=builders)
    assert any("logical content differs" in item for item in errors)


def test_verify_detects_stale_index(tmp_path: Path) -> None:
    root = tmp_path / "fixtures"
    spec = _spec("indexed")
    _materialize(root, [spec])
    (root / "index.json").write_text(
        json.dumps(
            {
                "schema_version": 1,
                "source_of_truth": "tests.generators.catalog.ALL_SPECS",
                "scenarios": ["indexed", "ghost"],
                "workbooks": [spec.relative_workbook],
                "count": 2,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    errors = verify_all(specs=[spec], root=root)
    assert any("index.json" in item for item in errors)
