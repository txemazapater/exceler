"""Phase 2B region-detection synthetic scenarios."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tests.generators.workbook_factory import ScenarioSpec, bold_header, new_workbook, write_matrix

THIN = Side(style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
BLOCK_A_FILL = PatternFill("solid", fgColor="FFF2CC")
BLOCK_B_FILL = PatternFill("solid", fgColor="E2EFDA")


def build_table_with_totals_footer(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Totals")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Item", "Qty", "Amount"],
            ["A", 1, 10],
            ["B", 2, 20],
            ["C", 3, 30],
            ["TOTAL", None, "=SUM(C2:C4)"],
        ],
    )
    bold_header(ws, 1, 3)
    for col in range(1, 4):
        ws.cell(1, col).fill = HEADER_FILL
    ws.cell(5, 1).font = Font(bold=True)
    ws.cell(5, 3).font = Font(bold=True)
    return wb


def build_note_block_below_table(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Notes")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Code", "Name"],
            ["1", "Alpha"],
            ["2", "Beta"],
        ],
    )
    bold_header(ws, 1, 2)
    ws["A6"] = "Observaciones sintéticas — no tabular"
    ws["A7"] = "Revisar manualmente antes de consolidar."
    return wb


def build_nested_title_and_table(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Nested")
    ws = wb.active
    assert ws is not None
    ws.merge_cells("A1:C1")
    ws["A1"] = "Catálogo sintético"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    write_matrix(
        ws,
        [
            ["Id", "Label", "Value"],
            [1, "Uno", 10],
            [2, "Dos", 20],
            [3, "Tres", 30],
        ],
        start_row=3,
    )
    bold_header(ws, 3, 3)
    for col in range(1, 4):
        ws.cell(3, col).fill = HEADER_FILL
    return wb


def build_false_gap_inside_table(_seed: int) -> Workbook:
    """One empty row inside an otherwise continuous table — must stay one region."""
    wb = new_workbook(sheet_title="Gap")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Sku", "Qty"],
            ["S-1", 1],
            ["S-2", 2],
        ],
    )
    # Row 4 intentionally empty; data continues at row 5.
    write_matrix(ws, [["S-3", 3], ["S-4", 4]], start_row=5)
    bold_header(ws, 1, 2)
    for col in range(1, 3):
        ws.cell(1, col).fill = HEADER_FILL
        for row in range(1, 7):
            if ws.cell(row, col).value is not None or row == 4:
                ws.cell(row, col).border = BOX
    return wb


def build_styled_separator_blocks(_seed: int) -> Workbook:
    """Two vertically stacked blocks with different fills and a blank separator row."""
    wb = new_workbook(sheet_title="Styled")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["A_Id", "A_Name"],
            [1, "North"],
            [2, "South"],
        ],
    )
    for row in range(1, 4):
        for col in range(1, 3):
            ws.cell(row, col).fill = BLOCK_A_FILL
            ws.cell(row, col).border = BOX
    # Row 4 blank separator
    write_matrix(
        ws,
        [
            ["B_Code", "B_Value"],
            ["X", 100],
            ["Y", 200],
        ],
        start_row=5,
    )
    for row in range(5, 8):
        for col in range(1, 3):
            ws.cell(row, col).fill = BLOCK_B_FILL
            ws.cell(row, col).border = BOX
    bold_header(ws, 1, 2)
    bold_header(ws, 5, 2)
    return wb


REGION_SPECS: list[ScenarioSpec] = [
    ScenarioSpec(
        scenario_id="table_with_totals_footer",
        category="structural",
        description="Tabla con fila de totales/fórmulas al pie",
        relative_workbook="workbooks/structural/table_with_totals_footer.xlsx",
        generator_name="region_scenarios.build_table_with_totals_footer",
        intentions=["Una región table", "Última fila candidata a footer"],
        features=["regions", "table", "footer"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {
                "sheets": [
                    {
                        "name": "Totals",
                        "region_count_min": 1,
                        "regions": [
                            {
                                "region_type": "table",
                                "bbox": {
                                    "first_row": 1,
                                    "last_row": 5,
                                    "first_col": 1,
                                    "last_col": 3,
                                },
                                "has_footer": True,
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="note_block_below_table",
        category="structural",
        description="Tabla seguida de bloque de notas separado",
        relative_workbook="workbooks/structural/note_block_below_table.xlsx",
        generator_name="region_scenarios.build_note_block_below_table",
        intentions=["Región table + región note/unknown"],
        features=["regions", "table", "note"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {
                "sheets": [
                    {
                        "name": "Notes",
                        "region_count_min": 2,
                        "contains_types": ["table", "note"],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="nested_title_and_table",
        category="structural",
        description="Título merge encima de tabla (jerarquía ligera)",
        relative_workbook="workbooks/structural/nested_title_and_table.xlsx",
        generator_name="region_scenarios.build_nested_title_and_table",
        intentions=["Título parent de tabla"],
        features=["regions", "title", "table", "nested"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {
                "sheets": [
                    {
                        "name": "Nested",
                        "region_count_min": 2,
                        "contains_types": ["title", "table"],
                        "has_parent_child": True,
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="false_gap_inside_table",
        category="structural",
        description="Fila vacía interior que no debe dividir la tabla",
        relative_workbook="workbooks/structural/false_gap_inside_table.xlsx",
        generator_name="region_scenarios.build_false_gap_inside_table",
        intentions=["Una sola región table pese a gap"],
        features=["regions", "table", "negative_split"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {
                "sheets": [
                    {
                        "name": "Gap",
                        "region_count": 1,
                        "regions": [
                            {
                                "region_type": "table",
                                "bbox": {
                                    "first_row": 1,
                                    "last_row": 6,
                                    "first_col": 1,
                                    "last_col": 2,
                                },
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="styled_separator_blocks",
        category="structural",
        description="Dos bloques con fills distintos separados por fila vacía",
        relative_workbook="workbooks/structural/styled_separator_blocks.xlsx",
        generator_name="region_scenarios.build_styled_separator_blocks",
        intentions=["Exactamente dos regiones"],
        features=["regions", "style_separator"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {
                "sheets": [
                    {
                        "name": "Styled",
                        "region_count": 2,
                    }
                ]
            },
        },
    ),
]


BUILDERS: dict[str, object] = {
    "region_scenarios.build_table_with_totals_footer": build_table_with_totals_footer,
    "region_scenarios.build_note_block_below_table": build_note_block_below_table,
    "region_scenarios.build_nested_title_and_table": build_nested_title_and_table,
    "region_scenarios.build_false_gap_inside_table": build_false_gap_inside_table,
    "region_scenarios.build_styled_separator_blocks": build_styled_separator_blocks,
}
