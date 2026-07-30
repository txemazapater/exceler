from __future__ import annotations

import json
import tempfile
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from tests.generators.catalog import ALL_SPECS, get_builder, get_special_saver
from tests.generators.workbook_factory import (
    EXPECTED_SCHEMA_VERSION,
    ScenarioSpec,
    expected_path,
    file_sha256,
    fixtures_root,
    index_payload,
    manifest_path,
    save_workbook,
    workbook_path,
)


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    return str(value)


def _cell_snapshot(ws: Worksheet) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    max_row = int(ws.max_row or 1)
    max_col = int(ws.max_column or 1)
    declared_area = max_row * max_col
    # Avoid walking pathological rectangles during fixture verification.
    if declared_area > 500_000:
        cells_map = getattr(ws, "_cells", {})
        iterable = [cell for _coord, cell in sorted(cells_map.items())]
    else:
        iterable = [
            cell
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col)
            for cell in row
        ]
    for cell in iterable:
        if cell.value is None and not cell.comment and not cell.hyperlink:
            # Keep empty cells only when they carry deliberate formatting used by fixtures.
            if cell.number_format in (None, "General") and not (cell.font and cell.font.bold):
                continue
        entry: dict[str, Any] = {
            "coord": cell.coordinate,
            "value": _json_safe(cell.value),
            "data_type": cell.data_type,
            "number_format": cell.number_format,
        }
        if isinstance(cell.value, str) and cell.value.startswith("="):
            entry["formula"] = cell.value
        elif cell.data_type == "f" and cell.value is not None:
            entry["formula"] = str(cell.value)
        if cell.comment is not None:
            entry["comment"] = cell.comment.text
        if cell.hyperlink is not None:
            entry["hyperlink"] = cell.hyperlink.target
        if cell.font is not None and cell.font.bold:
            entry["font_bold"] = True
        cells.append(entry)
    cells.sort(key=lambda item: item["coord"])
    return cells


def _column_dimensions(ws: Worksheet) -> dict[str, Any]:
    hidden: list[str] = []
    widths: dict[str, float] = {}
    for letter, dim in sorted(ws.column_dimensions.items()):
        if dim.hidden:
            hidden.append(letter)
        if dim.width is not None:
            widths[letter] = float(dim.width)
    return {"hidden": hidden, "widths": widths}


def _row_dimensions(ws: Worksheet) -> dict[str, Any]:
    hidden: list[int] = []
    heights: dict[str, float] = {}
    for index, dim in sorted(ws.row_dimensions.items(), key=lambda item: int(item[0])):
        if dim.hidden:
            hidden.append(int(index))
        if dim.height is not None:
            heights[str(int(index))] = float(dim.height)
    return {"hidden": hidden, "heights": heights}


def _table_snapshot(ws: Worksheet) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    for name in sorted(ws.tables.keys()):
        table = ws.tables[name]
        auto_filter = None
        if table.autoFilter is not None:
            auto_filter = {
                "ref": table.autoFilter.ref,
            }
        tables.append(
            {
                "name": name,
                "displayName": table.displayName,
                "ref": table.ref,
                "totalsRowCount": int(table.totalsRowCount or 0),
                "headerRowCount": int(table.headerRowCount or 0),
                "autoFilter": auto_filter,
            }
        )
    return tables


def _defined_names_snapshot(wb: Any) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    for name in sorted(wb.defined_names.keys()):
        defined = wb.defined_names[name]
        names.append(
            {
                "name": name,
                "attr_text": defined.attr_text,
                "localSheetId": defined.localSheetId,
            }
        )
    return names


def _external_links_snapshot(wb: Any) -> list[str]:
    links: list[str] = []
    external_links = getattr(wb, "_external_links", None) or []
    for link in external_links:
        target = getattr(link, "file_link", None)
        if target is not None and getattr(target, "Target", None):
            links.append(str(target.Target))
        else:
            links.append(str(link))
    return sorted(links)


def logical_snapshot(path: Path) -> dict[str, Any]:
    """Stable, JSON-serializable logical contract for a workbook fixture."""
    wb = load_workbook(path, data_only=False)
    sheets: list[dict[str, Any]] = []
    for index, name in enumerate(wb.sheetnames):
        ws = wb[name]
        auto_filter = None
        if ws.auto_filter is not None and ws.auto_filter.ref:
            auto_filter = {"ref": ws.auto_filter.ref}
        sheets.append(
            {
                "title": name,
                "index": index,
                "sheet_state": ws.sheet_state,
                "dimensions": ws.calculate_dimension(),
                "freeze_panes": ws.freeze_panes,
                "auto_filter": auto_filter,
                "column_dimensions": _column_dimensions(ws),
                "row_dimensions": _row_dimensions(ws),
                "merged": sorted(str(rng) for rng in ws.merged_cells.ranges),
                "tables": _table_snapshot(ws),
                "cells": _cell_snapshot(ws),
            }
        )
    return {
        "sheet_order": list(wb.sheetnames),
        "sheets": sheets,
        "defined_names": _defined_names_snapshot(wb),
        "external_links": _external_links_snapshot(wb),
    }


def _ensure_under_root(
    path: Path, root: Path, label: str, scenario_id: str, errors: list[str]
) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        errors.append(f"{scenario_id}: {label} escapes fixtures root ({path})")
        return False


def _validate_expected(
    spec: ScenarioSpec,
    *,
    root: Path,
    errors: list[str],
) -> None:
    exp = expected_path(spec, root=root)
    if not exp.exists():
        errors.append(f"{spec.scenario_id}: missing expected skeleton {exp.relative_to(root)}")
        return
    if not _ensure_under_root(exp, root, "expected", spec.scenario_id, errors):
        return
    try:
        payload = json.loads(exp.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        errors.append(f"{spec.scenario_id}: invalid expected JSON ({exc})")
        return
    if payload.get("schema_version") != EXPECTED_SCHEMA_VERSION:
        errors.append(
            f"{spec.scenario_id}: expected schema_version must be {EXPECTED_SCHEMA_VERSION}"
        )
    if payload.get("scenario_id") != spec.scenario_id:
        errors.append(f"{spec.scenario_id}: expected scenario_id mismatch")
    workbook_ref = spec.relative_workbook.replace("\\", "/")
    if payload.get("workbook") != workbook_ref:
        errors.append(f"{spec.scenario_id}: expected workbook path mismatch")
    expectations = payload.get("expectations")
    if not isinstance(expectations, dict):
        errors.append(f"{spec.scenario_id}: expected.expectations must be an object")
    elif expectations != dict(spec.expected_skeleton):
        errors.append(f"{spec.scenario_id}: expected.expectations diverge from catalog")


def _validate_index(*, root: Path, specs: list[ScenarioSpec], errors: list[str]) -> None:
    index_path = root / "index.json"
    if not index_path.exists():
        errors.append("missing fixtures index.json")
        return
    try:
        payload = json.loads(index_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        errors.append(f"invalid index.json ({exc})")
        return
    expected = index_payload(specs)
    if payload != expected:
        errors.append("index.json diverges from catalog (ALL_SPECS is the source of truth)")
    scenarios = payload.get("scenarios")
    if not isinstance(scenarios, list):
        errors.append("index.json scenarios must be a list")
        return
    if len(scenarios) != len(set(scenarios)):
        errors.append("index.json contains duplicate scenario ids")
    catalog_ids = [spec.scenario_id for spec in specs]
    missing = [sid for sid in catalog_ids if sid not in scenarios]
    extra = [sid for sid in scenarios if sid not in catalog_ids]
    for sid in missing:
        errors.append(f"index.json missing scenario: {sid}")
    for sid in extra:
        errors.append(f"index.json unknown scenario: {sid}")


def _has_vba_project(path: Path) -> bool:
    with zipfile.ZipFile(path) as archive:
        return any(name.lower().endswith("vbaproject.bin") for name in archive.namelist())


def verify_all(
    *,
    specs: list[ScenarioSpec] | None = None,
    root: Path | None = None,
    builders: dict[str, Any] | None = None,
) -> list[str]:
    """Validate manifests, expected skeletons, index and logical workbook contracts."""
    errors: list[str] = []
    active_specs = list(ALL_SPECS if specs is None else specs)
    active_root = fixtures_root() if root is None else root
    by_id = {spec.scenario_id: spec for spec in active_specs}
    if len(by_id) != len(active_specs):
        errors.append("Duplicate scenario_id values in catalog")

    seen_workbooks: set[str] = set()
    for spec in active_specs:
        if spec.seed is None:
            errors.append(f"{spec.scenario_id}: seed is missing")
        wb_key = spec.relative_workbook.replace("\\", "/")
        if wb_key in seen_workbooks:
            errors.append(f"Duplicate workbook path: {wb_key}")
        seen_workbooks.add(wb_key)

        man = manifest_path(spec, root=active_root)
        wb = workbook_path(spec, root=active_root)
        if not man.exists():
            errors.append(f"{spec.scenario_id}: missing manifest {man.relative_to(active_root)}")
            _validate_expected(spec, root=active_root, errors=errors)
            continue
        if not _ensure_under_root(man, active_root, "manifest", spec.scenario_id, errors):
            continue

        try:
            payload = json.loads(man.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            errors.append(f"{spec.scenario_id}: invalid manifest JSON ({exc})")
            _validate_expected(spec, root=active_root, errors=errors)
            continue

        if payload.get("scenario_id") != spec.scenario_id:
            errors.append(f"{spec.scenario_id}: manifest scenario_id mismatch")
        if payload.get("seed") != spec.seed:
            errors.append(f"{spec.scenario_id}: manifest seed mismatch")
        if payload.get("generator") != spec.generator_name:
            errors.append(f"{spec.scenario_id}: manifest generator mismatch")
        if payload.get("workbook") != wb_key:
            errors.append(f"{spec.scenario_id}: manifest workbook path mismatch")

        _validate_expected(spec, root=active_root, errors=errors)

        if not wb.exists():
            errors.append(f"{spec.scenario_id}: missing workbook {wb.relative_to(active_root)}")
            continue
        if not _ensure_under_root(wb, active_root, "workbook", spec.scenario_id, errors):
            continue

        try:
            load_workbook(wb, read_only=True, data_only=False)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{spec.scenario_id}: workbook is not readable ({exc})")
            continue

        if "no_vba_project" in spec.features and _has_vba_project(wb):
            errors.append(f"{spec.scenario_id}: unexpected vbaProject.bin present")

        if spec.versioned and spec.seed is not None:
            resolve_builder = (lambda name: builders[name]) if builders is not None else get_builder
            try:
                builder = resolve_builder(spec.generator_name)
            except KeyError:
                errors.append(f"{spec.scenario_id}: unknown generator {spec.generator_name}")
                continue
            regenerated = builder(spec.seed)
            with tempfile.TemporaryDirectory(prefix="exceler-fixture-") as tmp:
                tmp_root = Path(tmp)
                tmp_path = tmp_root / wb_key
                saver = None
                if builders is None:
                    saver = get_special_saver(spec.generator_name)
                if saver is not None:
                    saver(regenerated, tmp_path)
                else:
                    save_workbook(regenerated, tmp_path)
                # Optional diagnostic only — binary hash must not fail verification.
                _ = file_sha256(tmp_path)
                _ = file_sha256(wb)
                if logical_snapshot(tmp_path) != logical_snapshot(wb):
                    errors.append(
                        f"{spec.scenario_id}: regenerated logical content differs "
                        "from versioned workbook"
                    )

    workbooks_dir = active_root / "workbooks"
    if workbooks_dir.exists():
        expected_files = {workbook_path(spec, root=active_root).resolve() for spec in active_specs}
        for path in workbooks_dir.rglob("*"):
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}:
                if path.resolve() not in expected_files:
                    errors.append(f"Orphan workbook: {path.relative_to(active_root)}")

    manifests_dir = active_root / "manifests"
    if manifests_dir.exists():
        expected_manifests = {
            manifest_path(spec, root=active_root).resolve() for spec in active_specs
        }
        for path in manifests_dir.rglob("*.json"):
            if path.resolve() not in expected_manifests:
                errors.append(f"Orphan manifest: {path.relative_to(active_root)}")

    expected_dir = active_root / "expected"
    if expected_dir.exists():
        expected_files = {expected_path(spec, root=active_root).resolve() for spec in active_specs}
        seen_expected_ids: dict[str, Path] = {}
        for path in expected_dir.rglob("*.json"):
            if path.resolve() not in expected_files:
                errors.append(f"Orphan expected: {path.relative_to(active_root)}")
                continue
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                continue
            sid = payload.get("scenario_id")
            if isinstance(sid, str):
                if sid in seen_expected_ids:
                    errors.append(f"Duplicate expected scenario_id: {sid}")
                seen_expected_ids[sid] = path

    _validate_index(root=active_root, specs=active_specs, errors=errors)
    return errors


def main() -> None:
    errors = verify_all()
    if errors:
        print("Fixture verification FAILED:")
        for item in errors:
            print(f" - {item}")
        raise SystemExit(1)
    print(f"Fixture verification OK ({len(ALL_SPECS)} scenarios).")


if __name__ == "__main__":
    main()
