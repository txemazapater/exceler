from __future__ import annotations

import hashlib
import json
from collections.abc import Callable
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from tests.generators import DEFAULT_SEED, FIXTURES_ROOT

EXPECTED_SCHEMA_VERSION = 1
INDEX_SCHEMA_VERSION = 1


@dataclass(frozen=True)
class ScenarioSpec:
    scenario_id: str
    category: str
    description: str
    relative_workbook: str
    generator_name: str
    seed: int | None = DEFAULT_SEED
    intentions: list[str] = field(default_factory=list)
    features: list[str] = field(default_factory=list)
    versioned: bool = True
    expected_skeleton: dict[str, Any] = field(default_factory=dict)


Builder = Callable[[int], Workbook]


def fixtures_root() -> Path:
    return FIXTURES_ROOT


def workbook_path(spec: ScenarioSpec, *, root: Path | None = None) -> Path:
    base = fixtures_root() if root is None else root
    return base / spec.relative_workbook


def manifest_path(spec: ScenarioSpec, *, root: Path | None = None) -> Path:
    base = fixtures_root() if root is None else root
    return base / "manifests" / spec.category / f"{spec.scenario_id}.json"


def expected_path(spec: ScenarioSpec, *, root: Path | None = None) -> Path:
    base = fixtures_root() if root is None else root
    return base / "expected" / spec.category / f"{spec.scenario_id}.json"


def index_payload(specs: list[ScenarioSpec]) -> dict[str, Any]:
    """Derived index document. Python catalog ALL_SPECS remains the source of truth."""
    return {
        "schema_version": INDEX_SCHEMA_VERSION,
        "source_of_truth": "tests.generators.catalog.ALL_SPECS",
        "scenarios": [spec.scenario_id for spec in specs],
        "workbooks": [spec.relative_workbook.replace("\\", "/") for spec in specs],
        "count": len(specs),
    }


def new_workbook(*, sheet_title: str = "Sheet1") -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_title
    return wb


def write_matrix(
    ws: Any,
    rows: list[list[Any]],
    *,
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    for r_index, row in enumerate(rows):
        for c_index, value in enumerate(row):
            ws.cell(row=start_row + r_index, column=start_col + c_index, value=value)


def save_workbook(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_manifest(spec: ScenarioSpec, *, root: Path | None = None) -> None:
    path = manifest_path(spec, root=root)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "scenario_id": spec.scenario_id,
        "description": spec.description,
        "workbook": spec.relative_workbook.replace("\\", "/"),
        "generator": spec.generator_name,
        "seed": spec.seed,
        "intentions": list(spec.intentions),
        "features": list(spec.features),
        "versioned": spec.versioned,
    }
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_expected_skeleton(spec: ScenarioSpec, *, root: Path | None = None) -> None:
    path = expected_path(spec, root=root)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": EXPECTED_SCHEMA_VERSION,
        "scenario_id": spec.scenario_id,
        "workbook": spec.relative_workbook.replace("\\", "/"),
        "notes": "Skeleton for future Discovery Engine expectations. Not produced by analysis.",
        "expectations": dict(spec.expected_skeleton),
    }
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_index(specs: list[ScenarioSpec], *, root: Path | None = None) -> Path:
    base = fixtures_root() if root is None else root
    index_path = base / "index.json"
    index_path.write_text(
        json.dumps(index_payload(specs), indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return index_path


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_excel_table(ws: Any, *, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def bold_header(ws: Any, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        ws.cell(row=row, column=col).font = Font(bold=True)


def hide_column(ws: Any, column_index: int) -> None:
    ws.column_dimensions[get_column_letter(column_index)].hidden = True


def as_public_dict(spec: ScenarioSpec) -> dict[str, Any]:
    return asdict(spec)
