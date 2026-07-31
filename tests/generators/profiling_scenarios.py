"""Phase 2C profiling synthetic scenarios."""

from __future__ import annotations

from datetime import date, datetime, time

from openpyxl import Workbook
from openpyxl.styles import Font
from tests.generators.workbook_factory import ScenarioSpec, bold_header, new_workbook, write_matrix


def build_profile_core_types(_seed: int) -> Workbook:
    """Homogeneous physical/logical columns + blanks/null distinction."""
    wb = new_workbook(sheet_title="Core")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Code", "Amount", "Flag", "When", "At", "Clock", "Note"],
            [
                "0001",
                10.5,
                True,
                date(2026, 1, 1),
                datetime(2026, 1, 1, 12, 0, 0),
                time(9, 30),
                "alpha",
            ],
            [
                "0002",
                20.0,
                False,
                date(2026, 1, 2),
                datetime(2026, 1, 2, 13, 0, 0),
                time(10, 0),
                "beta",
            ],
            [
                "0003",
                30.25,
                True,
                date(2026, 1, 3),
                datetime(2026, 1, 3, 14, 0, 0),
                time(11, 15),
                "gamma",
            ],
        ],
    )
    bold_header(ws, 1, 7)
    for row, code in enumerate(["0001", "0002", "0003"], start=2):
        cell = ws.cell(row, 1, value=code)
        cell.number_format = "@"
        ws.cell(row, 2).number_format = "#,##0.00"
    return wb


def build_profile_logical_specials(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Specials")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Pct", "Money", "Uuid", "Email", "Url", "Phone", "Postal"],
            [
                0.15,
                125.5,
                "550e8400-e29b-41d4-a716-446655440000",
                "a@example.com",
                "https://example.com/a",
                "+34 600 111 222",
                "28001",
            ],
            [
                0.25,
                220.0,
                "123e4567-e89b-12d3-a456-426614174000",
                "b@example.org",
                "https://example.com/b",
                "+34 600 333 444",
                "08001",
            ],
            [
                "35%",
                "€80.00",
                "123e4567-e89b-12d3-a456-426614174001",
                "c@example.net",
                "http://example.com/c",
                "+34 600 555 666",
                "41001",
            ],
        ],
    )
    bold_header(ws, 1, 7)
    for row in range(2, 4):
        ws.cell(row, 1).number_format = "0%"
        ws.cell(row, 2).number_format = "€#,##0.00"
    return wb


def build_profile_mixed_and_anomalies(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Mixed")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["MaybeDate", "MaybeInt", "Blankish", "Err", "Formula", "Status"],
            ["2026-01-01", 1, "KEEP", "#DIV/0!", "=B2*2", "Pendiente"],
            ["03/04/2026", 2, "   ", "#N/A", "=B3*2", "Enviado"],
            ["error", "x", None, "#REF!", "=B4*2", "Pendiente"],
            ["TOTAL", None, None, None, "=SUM(B2:B4)", ""],
        ],
    )
    bold_header(ws, 1, 6)
    ws.cell(2, 3).value = ""
    ws.cell(5, 1).font = Font(bold=True)
    # Mark footer-like last row via bold — detector may or may not; expectations tolerate
    return wb


def build_profile_id_and_category(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Keys")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Id", "DupId", "NullableId", "Category", "FreeText"],
            ["A-1", "X", "N-1", "Activo", "lorem ipsum dolor sit amet"],
            ["A-2", "X", None, "Inactivo", "another long free text value here"],
            ["A-3", "Y", "N-3", "Activo", "yet another descriptive sentence"],
            ["A-4", "Y", "N-4", "Inactivo", "more free-form commentary text"],
        ],
    )
    bold_header(ws, 1, 5)
    return wb


def build_profile_headers(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Headers")
    ws = wb.active
    assert ws is not None
    # Sheet with title+table already covered elsewhere; here single header + no-header sheet
    write_matrix(
        ws,
        [
            ["Grupo", "Campo"],
            ["Ventas", "Unidades"],
            [10, 1],
            [20, 2],
        ],
    )
    # Multi-row header: rows 1-2 look header-like; detector may only mark row 1.
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    bare = wb.create_sheet("NoHeader")
    write_matrix(bare, [[1, "a"], [2, "b"], [3, "c"]])
    return wb


PROFILE_SPECS: list[ScenarioSpec] = [
    ScenarioSpec(
        scenario_id="profile_core_types",
        category="types",
        description="Columnas homogéneas: code/decimal/bool/date/datetime/time/text",
        relative_workbook="workbooks/types/profile_core_types.xlsx",
        generator_name="profiling_scenarios.build_profile_core_types",
        intentions=["Inferir tipos físicos y lógicos básicos", "Conservar ceros iniciales"],
        features=["profiling", "leading_zeroes", "dates"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {"sheets": [{"name": "Core", "region_count_min": 1}]},
            "profiling": {
                "sheets": [
                    {
                        "name": "Core",
                        "regions": [
                            {
                                "columns": [
                                    {
                                        "column_index": 1,
                                        "logical_type": "code",
                                        "minimum_confidence": 0.5,
                                        "identifier_candidate": True,
                                    },
                                    {
                                        "column_index": 2,
                                        "logical_type": "number",
                                        "minimum_confidence": 0.5,
                                    },
                                    {
                                        "column_index": 3,
                                        "logical_type": "boolean",
                                        "minimum_confidence": 0.5,
                                    },
                                ]
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="profile_logical_specials",
        category="types",
        description="Porcentaje, moneda, uuid, email, url, phone, postal",
        relative_workbook="workbooks/types/profile_logical_specials.xlsx",
        generator_name="profiling_scenarios.build_profile_logical_specials",
        intentions=["Tipos lógicos estructurales vía formato y patrón"],
        features=["profiling", "percentage", "currency", "uuid", "email"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "profiling": {
                "sheets": [
                    {
                        "name": "Specials",
                        "regions": [
                            {
                                "columns": [
                                    {
                                        "column_index": 1,
                                        "logical_type": "percentage",
                                        "minimum_confidence": 0.4,
                                    },
                                    {
                                        "column_index": 3,
                                        "logical_type": "uuid",
                                        "minimum_confidence": 0.5,
                                    },
                                    {
                                        "column_index": 4,
                                        "logical_type": "email",
                                        "minimum_confidence": 0.5,
                                    },
                                ]
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="profile_mixed_and_anomalies",
        category="types",
        description="Mezclas, fechas ambiguas, errores Excel, fórmulas, blank vs empty",
        relative_workbook="workbooks/types/profile_mixed_and_anomalies.xlsx",
        generator_name="profiling_scenarios.build_profile_mixed_and_anomalies",
        intentions=["Anomalías estructurales", "Fórmulas sin evaluar"],
        features=["profiling", "anomalies", "formulas", "ambiguous_dates"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "profiling": {
                "sheets": [
                    {
                        "name": "Mixed",
                        "regions": [
                            {
                                "columns": [
                                    {
                                        "column_index": 1,
                                        "logical_type": "date",
                                        "has_anomalies": True,
                                        "minimum_confidence": 0.3,
                                    },
                                    {
                                        "column_index": 4,
                                        "has_anomalies": True,
                                    },
                                    {
                                        "column_index": 5,
                                        "formula_count_min": 1,
                                    },
                                    {
                                        "column_index": 6,
                                        "categorical_candidate": True,
                                    },
                                ]
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="profile_id_and_category",
        category="types",
        description="Identificadores únicos/duplicados y categoría de baja cardinalidad",
        relative_workbook="workbooks/types/profile_id_and_category.xlsx",
        generator_name="profiling_scenarios.build_profile_id_and_category",
        intentions=["Candidatos a identifier y categorical"],
        features=["profiling", "identifier", "categorical"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "profiling": {
                "sheets": [
                    {
                        "name": "Keys",
                        "regions": [
                            {
                                "columns": [
                                    {
                                        "column_index": 1,
                                        "identifier_candidate": True,
                                        "minimum_confidence": 0.5,
                                    },
                                    {
                                        "column_index": 2,
                                        "identifier_candidate": False,
                                    },
                                    {
                                        "column_index": 4,
                                        "categorical_candidate": True,
                                    },
                                ]
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="profile_headers",
        category="types",
        description="Cabecera multi-fila y hoja sin cabecera",
        relative_workbook="workbooks/types/profile_headers.xlsx",
        generator_name="profiling_scenarios.build_profile_headers",
        intentions=["Separar headers de datos", "effective_name sin cabecera"],
        features=["profiling", "headers"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "profiling": {
                "sheets": [
                    {"name": "Headers", "region_count_min": 1},
                    {"name": "NoHeader", "region_count_min": 1},
                ]
            },
        },
    ),
]


BUILDERS: dict[str, object] = {
    "profiling_scenarios.build_profile_core_types": build_profile_core_types,
    "profiling_scenarios.build_profile_logical_specials": build_profile_logical_specials,
    "profiling_scenarios.build_profile_mixed_and_anomalies": build_profile_mixed_and_anomalies,
    "profiling_scenarios.build_profile_id_and_category": build_profile_id_and_category,
    "profiling_scenarios.build_profile_headers": build_profile_headers,
}
