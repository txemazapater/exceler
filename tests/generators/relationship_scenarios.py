"""Phase 2D relationship / key synthetic scenarios."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from tests.generators.workbook_factory import ScenarioSpec, bold_header, new_workbook, write_matrix


def build_rel_simple_primary_key(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerCode", "Name"],
            ["C-001", "Alpha"],
            ["C-002", "Beta"],
            ["C-003", "Gamma"],
            ["C-004", "Delta"],
        ],
    )
    bold_header(ws, 1, 2)
    for row in range(2, 6):
        ws.cell(row, 1).number_format = "@"
    return wb


def build_rel_duplicate_identifier(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Dup")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Code", "Label"],
            ["X", "one"],
            ["X", "two"],
            ["Y", "three"],
            ["Y", "four"],
        ],
    )
    bold_header(ws, 1, 2)
    return wb


def build_rel_composite_key(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Lines")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["OrderId", "LineNo", "Qty"],
            ["O1", 1, 10],
            ["O1", 2, 5],
            ["O2", 1, 7],
            ["O2", 2, 3],
        ],
    )
    bold_header(ws, 1, 3)
    return wb


def build_rel_customers_orders(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerCode", "Name"],
            ["C-001", "Alpha"],
            ["C-002", "Beta"],
            ["C-003", "Gamma"],
        ],
    )
    bold_header(ws, 1, 2)
    for row in range(2, 5):
        ws.cell(row, 1).number_format = "@"

    orders = wb.create_sheet("Orders")
    write_matrix(
        orders,
        [
            ["OrderId", "CustomerCode", "Amount"],
            ["O-1", "C-001", 100],
            ["O-2", "C-001", 50],
            ["O-3", "C-002", 80],
            ["O-4", "C-003", 20],
        ],
    )
    orders["A1"].font = Font(bold=True)
    orders["B1"].font = Font(bold=True)
    orders["C1"].font = Font(bold=True)
    for row in range(2, 6):
        orders.cell(row, 2).number_format = "@"
    return wb


def build_rel_invoice_header_lines(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Invoices")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["InvoiceId", "Customer"],
            ["INV-1", "A"],
            ["INV-2", "B"],
        ],
    )
    bold_header(ws, 1, 2)

    lines = wb.create_sheet("InvoiceLines")
    write_matrix(
        lines,
        [
            ["InvoiceId", "Line", "Product"],
            ["INV-1", 1, "P1"],
            ["INV-1", 2, "P2"],
            ["INV-2", 1, "P3"],
        ],
    )
    for cell in ("A1", "B1", "C1"):
        lines[cell].font = Font(bold=True)
    return wb


def build_rel_orphan_and_partial(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Parents")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["ParentId", "Label"],
            ["P1", "one"],
            ["P2", "two"],
        ],
    )
    bold_header(ws, 1, 2)

    child = wb.create_sheet("Children")
    write_matrix(
        child,
        [
            ["ChildId", "ParentId"],
            ["C1", "P1"],
            ["C2", "P2"],
            ["C3", "P9"],  # orphan
            ["C4", None],  # nullable
        ],
    )
    for cell in ("A1", "B1"):
        child[cell].font = Font(bold=True)
    return wb


def build_rel_bridge_table(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Students")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["StudentId", "Name"],
            ["S1", "Ann"],
            ["S2", "Bob"],
        ],
    )
    bold_header(ws, 1, 2)

    courses = wb.create_sheet("Courses")
    write_matrix(
        courses,
        [
            ["CourseId", "Title"],
            ["K1", "Math"],
            ["K2", "History"],
        ],
    )
    for cell in ("A1", "B1"):
        courses[cell].font = Font(bold=True)

    bridge = wb.create_sheet("Enrollment")
    write_matrix(
        bridge,
        [
            ["StudentId", "CourseId"],
            ["S1", "K1"],
            ["S1", "K2"],
            ["S2", "K1"],
        ],
    )
    for cell in ("A1", "B1"):
        bridge[cell].font = Font(bold=True)
    return wb


def build_rel_integer_unique_not_surrogate(_seed: int) -> Workbook:
    """Accidentally unique integers (RowNum/Qty-like) without FK reference evidence."""
    wb = new_workbook(sheet_title="Rows")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Qty", "Label"],
            [10, "a"],
            [5, "b"],
            [7, "c"],
            [3, "d"],
        ],
    )
    bold_header(ws, 1, 2)
    return wb


def build_rel_numeric_customer_id_fk(_seed: int) -> Workbook:
    """Numeric CustomerId PK accepted via Orders.CustomerId FK parent reference."""
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerId", "Name"],
            [1001, "Alpha"],
            [1002, "Beta"],
            [1003, "Gamma"],
        ],
    )
    bold_header(ws, 1, 2)

    orders = wb.create_sheet("Orders")
    write_matrix(
        orders,
        [
            ["OrderId", "CustomerId", "Amount"],
            ["O-1", 1001, 100],
            ["O-2", 1001, 50],
            ["O-3", 1002, 80],
            ["O-4", 1003, 20],
        ],
    )
    for cell in ("A1", "B1", "C1"):
        orders[cell].font = Font(bold=True)
    return wb


def build_rel_matching_measures_no_relation(_seed: int) -> Workbook:
    """Identical unique measure domains must not invent PK/FK either direction."""
    wb = new_workbook(sheet_title="WarehouseA")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Qty", "Note"],
            [10, "a"],
            [20, "b"],
            [30, "c"],
            [40, "d"],
        ],
    )
    bold_header(ws, 1, 2)

    other = wb.create_sheet("WarehouseB")
    write_matrix(
        other,
        [
            ["Amount", "Note"],
            [10, "w"],
            [20, "x"],
            [30, "y"],
            [40, "z"],
        ],
    )
    for cell in ("A1", "B1"):
        other[cell].font = Font(bold=True)
    return wb


def build_rel_measure_into_identifier_no_fk(_seed: int) -> Workbook:
    """Sales.Amount ⊆ Customers.CustomerId must not invent an FK; Orders.CustomerId must."""
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerId", "Name"],
            [100, "Alpha"],
            [200, "Beta"],
            [300, "Gamma"],
            [400, "Delta"],
        ],
    )
    bold_header(ws, 1, 2)

    sales = wb.create_sheet("Sales")
    write_matrix(
        sales,
        [
            ["SaleId", "Amount"],
            ["S-1", 100],
            ["S-2", 200],
            ["S-3", 300],
            ["S-4", 150],
        ],
    )
    for cell in ("A1", "B1"):
        sales[cell].font = Font(bold=True)

    orders = wb.create_sheet("Orders")
    write_matrix(
        orders,
        [
            ["OrderId", "CustomerId", "Amount"],
            ["O-1", 100, 50],
            ["O-2", 100, 75],
            ["O-3", 200, 90],
            ["O-4", 300, 40],
        ],
    )
    for cell in ("A1", "B1", "C1"):
        orders[cell].font = Font(bold=True)
    return wb


def build_rel_incompatible_product_customer(_seed: int) -> Workbook:
    """ProductId ⊆ CustomerId must not invent an FK (incompatible entities)."""
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerId", "Name"],
            [100, "Alpha"],
            [200, "Beta"],
            [300, "Gamma"],
        ],
    )
    bold_header(ws, 1, 2)

    products = wb.create_sheet("Products")
    write_matrix(
        products,
        [
            ["RowId", "ProductId"],
            ["R1", 100],
            ["R2", 200],
            ["R3", 300],
        ],
    )
    for cell in ("A1", "B1"):
        products[cell].font = Font(bold=True)
    return wb


def build_rel_alias_client_customer(_seed: int) -> Workbook:
    """ClientId aliases to customer and references Customers.CustomerId."""
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerId", "Name"],
            [1001, "Alpha"],
            [1002, "Beta"],
            [1003, "Gamma"],
        ],
    )
    bold_header(ws, 1, 2)

    orders = wb.create_sheet("Orders")
    write_matrix(
        orders,
        [
            ["OrderId", "ClientId"],
            [5001, 1001],
            [5002, 1001],
            [5003, 1002],
        ],
    )
    for cell in ("A1", "B1"):
        orders[cell].font = Font(bold=True)
    return wb


def build_rel_alias_article_product(_seed: int) -> Workbook:
    """ArticleCode aliases to product and references Products.ProductCode."""
    wb = new_workbook(sheet_title="Products")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["ProductCode", "Name"],
            ["P-01", "Widget"],
            ["P-02", "Gadget"],
            ["P-03", "Doohickey"],
        ],
    )
    bold_header(ws, 1, 2)
    for row in range(2, 5):
        ws.cell(row, 1).number_format = "@"

    lines = wb.create_sheet("OrderLines")
    write_matrix(
        lines,
        [
            ["LineId", "ArticleCode", "Qty"],
            ["L1", "P-01", 2],
            ["L2", "P-01", 1],
            ["L3", "P-02", 4],
            ["L4", "P-03", 1],
        ],
    )
    for cell in ("A1", "B1", "C1"):
        lines[cell].font = Font(bold=True)
    for row in range(2, 6):
        lines.cell(row, 2).number_format = "@"
    return wb


def build_rel_insufficient_bare_id_fk(_seed: int) -> Workbook:
    """Bare Id overlapping CustomerId must not invent an FK (insufficient entity)."""
    wb = new_workbook(sheet_title="Customers")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["CustomerId", "Name"],
            [100, "Alpha"],
            [200, "Beta"],
            [300, "Gamma"],
        ],
    )
    bold_header(ws, 1, 2)

    source = wb.create_sheet("Source")
    write_matrix(
        source,
        [
            ["Id", "Note"],
            [100, "a"],
            [200, "b"],
            [100, "c"],
        ],
    )
    for cell in ("A1", "B1"):
        source[cell].font = Font(bold=True)
    return wb


def build_rel_pk_ranking(_seed: int) -> Workbook:
    """Code + unique free-text names: code must rank first among accepted PKs."""
    wb = new_workbook(sheet_title="People")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["PersonCode", "FullName"],
            ["P-01", "Alice Example"],
            ["P-02", "Bob Example"],
            ["P-03", "Carol Example"],
            ["P-04", "Dan Example"],
        ],
    )
    bold_header(ws, 1, 2)
    for row in range(2, 6):
        ws.cell(row, 1).number_format = "@"
    return wb


RELATIONSHIP_SPECS: list[ScenarioSpec] = [
    ScenarioSpec(
        scenario_id="rel_simple_primary_key",
        category="relationships",
        description="Unique non-null code as PK candidate",
        relative_workbook="workbooks/relationships/rel_simple_primary_key.xlsx",
        generator_name="relationship_scenarios.build_rel_simple_primary_key",
        intentions=["Detect primary key candidate"],
        features=["relationships", "primary_key"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Customers",
                        "pk_rank_order": [1],
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": True,
                                "key_kind": "natural",
                                "minimum_score": 0.45,
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_duplicate_identifier",
        category="relationships",
        description="Duplicated codes should not be strong PK",
        relative_workbook="workbooks/relationships/rel_duplicate_identifier.xlsx",
        generator_name="relationship_scenarios.build_rel_duplicate_identifier",
        intentions=["Reject duplicate identifier as PK"],
        features=["relationships", "false_pk"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Dup",
                        "accepted_primary_keys_max": 0,
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": False,
                                "rejection_reason": "below_min_pk_distinct_ratio",
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_composite_key",
        category="relationships",
        description="OrderId+LineNo composite uniqueness",
        relative_workbook="workbooks/relationships/rel_composite_key.xlsx",
        generator_name="relationship_scenarios.build_rel_composite_key",
        intentions=["Detect composite key"],
        features=["relationships", "composite_key"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Lines",
                        "composite_keys_min": 1,
                        "primary_keys": [
                            {
                                "column_index": 3,
                                "accepted": False,
                                "rejection_reason": "insufficient_independent_identifier_evidence",
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_customers_orders",
        category="relationships",
        description="Orders.CustomerCode references Customers.CustomerCode",
        relative_workbook="workbooks/relationships/rel_customers_orders.xlsx",
        generator_name="relationship_scenarios.build_rel_customers_orders",
        intentions=["Detect FK inclusion"],
        features=["relationships", "foreign_key"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "foreign_keys": [
                    {
                        "from_sheet": "Orders",
                        "from_column_index": 2,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": True,
                        "minimum_inclusion": 0.99,
                        "minimum_confidence": 0.35,
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_invoice_header_lines",
        category="relationships",
        description="Invoice header to lines 1:N",
        relative_workbook="workbooks/relationships/rel_invoice_header_lines.xlsx",
        generator_name="relationship_scenarios.build_rel_invoice_header_lines",
        intentions=["1:N cardinality"],
        features=["relationships", "one_to_many"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "foreign_keys": [
                    {
                        "from_sheet": "InvoiceLines",
                        "from_column_index": 1,
                        "to_sheet": "Invoices",
                        "to_column_index": 1,
                        "accepted": True,
                        "cardinality": "one_to_many",
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_orphan_and_partial",
        category="relationships",
        description="Orphan and nullable FK values",
        relative_workbook="workbooks/relationships/rel_orphan_and_partial.xlsx",
        generator_name="relationship_scenarios.build_rel_orphan_and_partial",
        intentions=["Surface orphans and uncertainty"],
        features=["relationships", "orphans"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "foreign_keys": [
                    {
                        "from_sheet": "Children",
                        "from_column_index": 2,
                        "to_sheet": "Parents",
                        "to_column_index": 1,
                        "has_orphans": True,
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_bridge_table",
        category="relationships",
        description="Enrollment bridge between students and courses",
        relative_workbook="workbooks/relationships/rel_bridge_table.xlsx",
        generator_name="relationship_scenarios.build_rel_bridge_table",
        intentions=["Structural N:M / bridge"],
        features=["relationships", "many_to_many"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 3}},
            "relationships": {
                "foreign_keys_min": 1,
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_integer_unique_not_surrogate",
        category="relationships",
        description="Accidentally unique Qty is rejected; never SURROGATE",
        relative_workbook="workbooks/relationships/rel_integer_unique_not_surrogate.xlsx",
        generator_name="relationship_scenarios.build_rel_integer_unique_not_surrogate",
        intentions=["INTEGER+unique without structural evidence is rejected"],
        features=["relationships", "key_kind", "false_pk"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Rows",
                        "accepted_primary_keys_max": 0,
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": False,
                                "key_kind": "primary",
                                "key_kind_not": "surrogate",
                                "rejection_reason": "insufficient_independent_identifier_evidence",
                            }
                        ],
                    }
                ]
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_numeric_customer_id_fk",
        category="relationships",
        description="Numeric CustomerId accepted with independent Id evidence + FK",
        relative_workbook="workbooks/relationships/rel_numeric_customer_id_fk.xlsx",
        generator_name="relationship_scenarios.build_rel_numeric_customer_id_fk",
        intentions=["Numeric PK with independent identity evidence"],
        features=["relationships", "numeric_identifier", "foreign_key"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Customers",
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": True,
                                "key_kind": "primary",
                                "key_kind_not": "surrogate",
                                "minimum_score": 0.45,
                            }
                        ],
                    },
                    {
                        "name": "Orders",
                        "primary_keys": [
                            {
                                "column_index": 3,
                                "accepted": False,
                                "rejection_reason": "insufficient_independent_identifier_evidence",
                            }
                        ],
                    },
                ],
                "foreign_keys": [
                    {
                        "from_sheet": "Orders",
                        "from_column_index": 2,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": True,
                        "minimum_inclusion": 0.99,
                    }
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_matching_measures_no_relation",
        category="relationships",
        description="Matching Qty/Amount domains invent neither PK nor FK",
        relative_workbook="workbooks/relationships/rel_matching_measures_no_relation.xlsx",
        generator_name="relationship_scenarios.build_rel_matching_measures_no_relation",
        intentions=["Break circular measure coincidence"],
        features=["relationships", "false_fk", "false_pk"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "sheets": [
                    {
                        "name": "WarehouseA",
                        "accepted_primary_keys_max": 0,
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": False,
                                "rejection_reason": "insufficient_independent_identifier_evidence",
                            }
                        ],
                    },
                    {
                        "name": "WarehouseB",
                        "accepted_primary_keys_max": 0,
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": False,
                                "rejection_reason": "insufficient_independent_identifier_evidence",
                            }
                        ],
                    },
                ],
                "foreign_keys_accepted_max": 0,
                "foreign_keys": [
                    {
                        "from_sheet": "WarehouseA",
                        "from_column_index": 1,
                        "to_sheet": "WarehouseB",
                        "to_column_index": 1,
                        "accepted": False,
                    },
                    {
                        "from_sheet": "WarehouseB",
                        "from_column_index": 1,
                        "to_sheet": "WarehouseA",
                        "to_column_index": 1,
                        "accepted": False,
                    },
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_measure_into_identifier_no_fk",
        category="relationships",
        description="Amount into CustomerId rejected; Orders.CustomerId FK preserved",
        relative_workbook="workbooks/relationships/rel_measure_into_identifier_no_fk.xlsx",
        generator_name="relationship_scenarios.build_rel_measure_into_identifier_no_fk",
        intentions=["Child reference evidence for FK sources"],
        features=["relationships", "false_fk", "child_reference"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 3}},
            "relationships": {
                "foreign_keys": [
                    {
                        "from_sheet": "Orders",
                        "from_column_index": 2,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": True,
                        "minimum_inclusion": 0.99,
                    },
                    {
                        "from_sheet": "Sales",
                        "from_column_index": 2,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": False,
                        "rejection_reason": "insufficient_child_reference_evidence",
                    },
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_incompatible_product_customer",
        category="relationships",
        description="ProductId into CustomerId rejected for incompatible entities",
        relative_workbook="workbooks/relationships/rel_incompatible_product_customer.xlsx",
        generator_name="relationship_scenarios.build_rel_incompatible_product_customer",
        intentions=["Reject cross-entity identifier value coincidence"],
        features=["relationships", "false_fk", "semantic_entity"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Customers",
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": True,
                                "key_kind": "primary",
                            }
                        ],
                    }
                ],
                "foreign_keys_accepted_max": 0,
                "foreign_keys": [
                    {
                        "from_sheet": "Products",
                        "from_column_index": 2,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": False,
                        "rejection_reason": "incompatible_reference_target_semantics",
                    },
                    {
                        "from_sheet": "Customers",
                        "from_column_index": 1,
                        "to_sheet": "Products",
                        "to_column_index": 2,
                        "accepted": False,
                        "rejection_reason": "incompatible_reference_target_semantics",
                    },
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_alias_client_customer",
        category="relationships",
        description="ClientId aliases to customer and references CustomerId",
        relative_workbook="workbooks/relationships/rel_alias_client_customer.xlsx",
        generator_name="relationship_scenarios.build_rel_alias_client_customer",
        intentions=["Accept declared customer/client alias"],
        features=["relationships", "foreign_key", "semantic_alias"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "sheets": [
                    {
                        "name": "Customers",
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": True,
                                "key_kind": "primary",
                            }
                        ],
                    }
                ],
                "foreign_keys": [
                    {
                        "from_sheet": "Orders",
                        "from_column_index": 2,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": True,
                        "minimum_inclusion": 0.99,
                    }
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_alias_article_product",
        category="relationships",
        description="ArticleCode aliases to product and references ProductCode",
        relative_workbook="workbooks/relationships/rel_alias_article_product.xlsx",
        generator_name="relationship_scenarios.build_rel_alias_article_product",
        intentions=["Accept declared product/article alias"],
        features=["relationships", "foreign_key", "semantic_alias"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "foreign_keys": [
                    {
                        "from_sheet": "OrderLines",
                        "from_column_index": 2,
                        "to_sheet": "Products",
                        "to_column_index": 1,
                        "accepted": True,
                        "minimum_inclusion": 0.99,
                    }
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_insufficient_bare_id_fk",
        category="relationships",
        description="Bare Id into CustomerId rejected for insufficient entity evidence",
        relative_workbook="workbooks/relationships/rel_insufficient_bare_id_fk.xlsx",
        generator_name="relationship_scenarios.build_rel_insufficient_bare_id_fk",
        intentions=["Reject bare structural Id without entity token"],
        features=["relationships", "false_fk", "semantic_insufficient"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 2}},
            "relationships": {
                "foreign_keys_accepted_max": 0,
                "foreign_keys": [
                    {
                        "from_sheet": "Source",
                        "from_column_index": 1,
                        "to_sheet": "Customers",
                        "to_column_index": 1,
                        "accepted": False,
                        "rejection_reason": "insufficient_reference_target_semantics",
                    }
                ],
            },
        },
    ),
    ScenarioSpec(
        scenario_id="rel_pk_ranking",
        category="relationships",
        description="Code ranks above unique free-text name",
        relative_workbook="workbooks/relationships/rel_pk_ranking.xlsx",
        generator_name="relationship_scenarios.build_rel_pk_ranking",
        intentions=["PK ranking prefers code/identifier over text"],
        features=["relationships", "ranking"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "relationships": {
                "sheets": [
                    {
                        "name": "People",
                        "pk_rank_order": [1],
                        "primary_keys": [
                            {
                                "column_index": 1,
                                "accepted": True,
                                "key_kind": "natural",
                            },
                            {
                                "column_index": 2,
                                "accepted": False,
                            },
                        ],
                    }
                ]
            },
        },
    ),
]

BUILDERS = {
    "relationship_scenarios.build_rel_simple_primary_key": build_rel_simple_primary_key,
    "relationship_scenarios.build_rel_duplicate_identifier": build_rel_duplicate_identifier,
    "relationship_scenarios.build_rel_composite_key": build_rel_composite_key,
    "relationship_scenarios.build_rel_customers_orders": build_rel_customers_orders,
    "relationship_scenarios.build_rel_invoice_header_lines": build_rel_invoice_header_lines,
    "relationship_scenarios.build_rel_orphan_and_partial": build_rel_orphan_and_partial,
    "relationship_scenarios.build_rel_bridge_table": build_rel_bridge_table,
    "relationship_scenarios.build_rel_integer_unique_not_surrogate": (
        build_rel_integer_unique_not_surrogate
    ),
    "relationship_scenarios.build_rel_numeric_customer_id_fk": build_rel_numeric_customer_id_fk,
    "relationship_scenarios.build_rel_matching_measures_no_relation": (
        build_rel_matching_measures_no_relation
    ),
    "relationship_scenarios.build_rel_measure_into_identifier_no_fk": (
        build_rel_measure_into_identifier_no_fk
    ),
    "relationship_scenarios.build_rel_incompatible_product_customer": (
        build_rel_incompatible_product_customer
    ),
    "relationship_scenarios.build_rel_alias_client_customer": build_rel_alias_client_customer,
    "relationship_scenarios.build_rel_alias_article_product": build_rel_alias_article_product,
    "relationship_scenarios.build_rel_insufficient_bare_id_fk": (build_rel_insufficient_bare_id_fk),
    "relationship_scenarios.build_rel_pk_ranking": build_rel_pk_ranking,
}
