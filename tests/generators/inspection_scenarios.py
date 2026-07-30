"""Phase 2A inspection-focused synthetic scenarios."""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from tests.generators.workbook_factory import (
    ScenarioSpec,
    bold_header,
    new_workbook,
    write_matrix,
)


def build_very_hidden_sheet(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Shown")
    ws = wb.active
    assert ws is not None
    write_matrix(ws, [["A"], [1]])
    hidden = wb.create_sheet("DeepHidden")
    write_matrix(hidden, [["X"], [9]])
    hidden.sheet_state = "veryHidden"
    return wb


def build_freeze_and_autofilter(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Frozen")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Code", "Qty", "Note"],
            ["A", 1, "one"],
            ["B", 2, "two"],
            ["C", 3, "three"],
        ],
    )
    bold_header(ws, 1, 3)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C4"
    return wb


def build_hidden_rows(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Rows")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Id", "Name"],
            [1, "keep"],
            [2, "hidden-row"],
            [3, "keep"],
        ],
    )
    ws.row_dimensions[3].hidden = True
    ws.row_dimensions[3].height = 30.0
    ws.column_dimensions["A"].width = 12.0
    return wb


def build_inflated_dimension(_seed: int) -> Workbook:
    """Dimension inflated by residual formatting without meaningful values far away."""
    wb = new_workbook(sheet_title="Inflated")
    ws = wb.active
    assert ws is not None
    write_matrix(ws, [["Id", "Name"], [1, "Alpha"]])
    far = ws.cell(row=500, column=20)
    far.number_format = "0.00"
    # No value — openpyxl still tracks the cell; declared dimension grows.
    return wb


def build_cell_physical_types(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Types")
    ws = wb.active
    assert ws is not None
    rows: list[list[Any]] = [
        ["Kind", "Value"],
        ["text", "hello"],
        ["integer", 42],
        ["decimal", Decimal("3.14")],
        ["boolean", True],
        ["date", date(2024, 1, 15)],
        ["datetime", datetime(2024, 1, 15, 13, 45, 0)],
        ["time", time(9, 30, 0)],
        ["percent", 0.25],
        ["currency", 12.5],
        ["numeric_as_text", "00123"],
        ["leading_zeros", "007"],
        ["empty_string", ""],
        ["blank", None],
        ["formula", None],
        ["error", None],
        ["comment_cell", "noted"],
        ["hyperlink_ext", "docs"],
        ["newline", "line1\nline2"],
        ["unicode", "café — 東京"],
        ["whitespace", "  padded  "],
    ]
    write_matrix(ws, rows)
    bold_header(ws, 1, 2)
    ws["B9"].number_format = "0%"
    ws["B10"].number_format = '"$"#,##0.00'
    ws["B15"] = "=1+1"
    ws["B16"] = "=1/0"
    ws["B17"].comment = Comment("synthetic comment", "exceler")
    ws["B18"].hyperlink = "https://example.invalid/docs"
    return wb


def build_defined_names_variants(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Names")
    ws = wb.active
    assert ws is not None
    write_matrix(ws, [["Label", "Amount"], ["A", 10], ["B", 20]])
    other = wb.create_sheet("Other")
    write_matrix(other, [["X"], [1]])
    wb.defined_names.add(DefinedName(name="GlobalRange", attr_text="'Names'!$A$1:$B$3"))
    wb.defined_names.add(DefinedName(name="GlobalConst", attr_text='"ACME"'))
    # Note: openpyxl does not reliably round-trip localSheetId-scoped names on save.
    # Local names are documented as a known limitation (docs/limitations.md).
    return wb


def build_merged_variants(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Merges")
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Horizontal title"
    ws.merge_cells("A1:C1")
    ws["A3"] = "Vertical"
    ws.merge_cells("A3:A5")
    ws.merge_cells("E1:F1")  # empty merge (no value)
    ws["A7"] = "Block"
    ws.merge_cells("A7:B8")
    write_matrix(ws, [["H1", "H2"], [1, 2]], start_row=10)
    return wb


def build_structured_tables_advanced(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Tables")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Sku", "Qty", "Total"],
            ["S-1", 1, 1.0],
            ["S-2", 2, 2.0],
            ["S-3", 3, 3.0],
            ["", "", 6.0],
        ],
        start_row=3,
        start_col=2,
    )
    table = Table(displayName="OffsetInventory", ref="B3:D7")
    table.totalsRowCount = 1
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    write_matrix(
        ws,
        [
            ["Region", "Count"],
            ["N", 1],
            ["S", 2],
        ],
        start_row=3,
        start_col=6,
    )
    table2 = Table(displayName="RegionsTable", ref="F3:G5")
    table2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table2)
    return wb


def build_cross_sheet_formula(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Inputs")
    ws = wb.active
    assert ws is not None
    write_matrix(ws, [["Value"], [10], [20]])
    calc = wb.create_sheet("Calc")
    calc["A1"] = "Sum"
    calc["B1"] = "=Inputs!A2+Inputs!A3"
    calc["B2"] = "=SUM(Inputs!A2:A3)"
    return wb


def build_named_range_formula(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Data")
    ws = wb.active
    assert ws is not None
    write_matrix(ws, [["N"], [1], [2], [3]])
    wb.defined_names.add(DefinedName(name="Nums", attr_text="'Data'!$A$2:$A$4"))
    ws["C1"] = "=SUM(Nums)"
    return wb


def _inject_vba_project(xlsx_bytes: bytes) -> bytes:
    """Embed a non-executable stub vbaProject.bin for presence detection only."""
    src = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "[Content_Types].xml":
                root = ET.fromstring(data)
                ns = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}
                exists = any(
                    el.get("PartName", "").lower() == "/xl/vbaproject.bin"
                    for el in root.findall("ct:Override", ns)
                )
                if not exists:
                    override = ET.SubElement(
                        root,
                        "{http://schemas.openxmlformats.org/package/2006/content-types}Override",
                    )
                    override.set("PartName", "/xl/vbaProject.bin")
                    override.set(
                        "ContentType",
                        "application/vnd.ms-office.vbaProject",
                    )
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            elif item.filename == "xl/_rels/workbook.xml.rels":
                root = ET.fromstring(data)
                rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
                exists = any(
                    el.get("Target", "").lower() == "vbaproject.bin"
                    for el in root.findall(f"{{{rel_ns}}}Relationship")
                )
                if not exists:
                    rel = ET.SubElement(root, f"{{{rel_ns}}}Relationship")
                    rel.set("Id", "rIdVba")
                    rel.set(
                        "Type",
                        "http://schemas.microsoft.com/office/2006/relationships/vbaProject",
                    )
                    rel.set("Target", "vbaProject.bin")
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            zout.writestr(item, data)
        # Inert stub — never executed; presence-only contract for Phase 2A.
        zout.writestr("xl/vbaProject.bin", b"EXCELER_VBA_STUB_NOT_EXECUTABLE\x00")
    return out.getvalue()


def build_xlsm_with_vba_stub(_seed: int) -> Workbook:
    """Return a workbook object; bytes with VBA stub are produced by save hook in catalog."""
    wb = new_workbook(sheet_title="WithVba")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Note", "Detail"],
            ["vba_stub_present", "has_vba_project_true_not_executed"],
        ],
    )
    return wb


def save_xlsm_with_vba_stub(wb: Workbook, path: Any) -> None:
    from pathlib import Path

    path = Path(path)
    buf = io.BytesIO()
    wb.save(buf)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_inject_vba_project(buf.getvalue()))


INSPECTION_SPECS: list[ScenarioSpec] = [
    ScenarioSpec(
        scenario_id="very_hidden_sheet",
        category="structural",
        description="Hoja veryHidden junto a una visible",
        relative_workbook="workbooks/structural/very_hidden_sheet.xlsx",
        generator_name="inspection_scenarios.build_very_hidden_sheet",
        intentions=["Una hoja visible y una veryHidden"],
        features=["very_hidden_sheet"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "has_vba_project": False, "worksheet_count": 2},
                "worksheets": [
                    {"name": "Shown", "index": 0, "visibility": "visible"},
                    {"name": "DeepHidden", "index": 1, "visibility": "veryHidden"},
                ],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="freeze_and_autofilter",
        category="structural",
        description="Freeze panes y autofiltro",
        relative_workbook="workbooks/structural/freeze_and_autofilter.xlsx",
        generator_name="inspection_scenarios.build_freeze_and_autofilter",
        intentions=["freeze_panes=A2", "auto_filter A1:C4"],
        features=["freeze_panes", "auto_filter"],
        expected_skeleton={
            "inspection": {
                "worksheets": [
                    {
                        "name": "Frozen",
                        "freeze_panes": "A2",
                        "auto_filter": "A1:C4",
                    }
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="hidden_rows",
        category="structural",
        description="Fila oculta con altura explícita",
        relative_workbook="workbooks/structural/hidden_rows.xlsx",
        generator_name="inspection_scenarios.build_hidden_rows",
        intentions=["Fila 3 oculta"],
        features=["hidden_rows"],
        expected_skeleton={
            "inspection": {
                "worksheets": [
                    {
                        "name": "Rows",
                        "hidden_rows": [3],
                    }
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="inflated_dimension",
        category="structural",
        description="Dimensión declarada inflada por formato residual",
        relative_workbook="workbooks/structural/inflated_dimension.xlsx",
        generator_name="inspection_scenarios.build_inflated_dimension",
        intentions=[
            "declared_dimension supera las celdas con valor",
            "No asumir max_row×max_column como datos reales",
        ],
        features=["inflated_dimension"],
        expected_skeleton={
            "inspection": {
                "worksheets": [
                    {
                        "name": "Inflated",
                        "declared_dimension_inflated": True,
                        "min_observed_cells": 3,
                    }
                ],
                "warnings_contain": ["DIMENSION_MAY_BE_INFLATED"],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="cell_physical_types",
        category="types",
        description="Tipos físicos de celda sin inferencia semántica",
        relative_workbook="workbooks/types/cell_physical_types.xlsx",
        generator_name="inspection_scenarios.build_cell_physical_types",
        intentions=["Conservar value/data_type/number_format/formula/comentario/hipervínculo"],
        features=["cell_types", "comments", "hyperlinks", "formulas"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "has_vba_project": False},
                "cells": [
                    {"sheet": "Types", "coordinate": "B15", "formula": "=1+1"},
                    {"sheet": "Types", "coordinate": "B16", "formula": "=1/0"},
                    {"sheet": "Types", "coordinate": "B3", "value_kind": "integer"},
                    {
                        "sheet": "Types",
                        "coordinate": "B11",
                        "value_kind": "string",
                        "text": "00123",
                    },
                    {"sheet": "Types", "coordinate": "B17", "has_comment": True},
                    {"sheet": "Types", "coordinate": "B18", "has_hyperlink": True},
                ],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="defined_names_variants",
        category="structural",
        description="Nombres definidos globales, locales y constantes",
        relative_workbook="workbooks/structural/defined_names_variants.xlsx",
        generator_name="inspection_scenarios.build_defined_names_variants",
        intentions=["GlobalRange (rango)", "GlobalConst (constante)"],
        features=["defined_names"],
        expected_skeleton={
            "inspection": {
                "defined_names": [
                    {"name": "GlobalConst"},
                    {"name": "GlobalRange"},
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="merged_variants",
        category="structural",
        description="Varios rangos combinados horizontales y verticales",
        relative_workbook="workbooks/structural/merged_variants.xlsx",
        generator_name="inspection_scenarios.build_merged_variants",
        intentions=["Merges A1:C1, A3:A5, E1:F1, A7:B8"],
        features=["merged_cells"],
        expected_skeleton={
            "inspection": {
                "worksheets": [
                    {
                        "name": "Merges",
                        "merged_ranges": ["A1:C1", "A3:A5", "A7:B8", "E1:F1"],
                    }
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="structured_tables_advanced",
        category="structural",
        description="Varias tablas estructuradas, una con totales y fuera de A1",
        relative_workbook="workbooks/structural/structured_tables_advanced.xlsx",
        generator_name="inspection_scenarios.build_structured_tables_advanced",
        intentions=["OffsetInventory con totalsRowCount=1", "RegionsTable"],
        features=["excel_table", "totals_row"],
        expected_skeleton={
            "inspection": {
                "worksheets": [
                    {
                        "name": "Tables",
                        "tables": [
                            {
                                "name": "OffsetInventory",
                                "ref": "B3:D7",
                                "totals_row_count": 1,
                            },
                            {"name": "RegionsTable", "ref": "F3:G5", "totals_row_count": 0},
                        ],
                    }
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="cross_sheet_formula",
        category="types",
        description="Fórmulas con referencia a otra hoja",
        relative_workbook="workbooks/types/cross_sheet_formula.xlsx",
        generator_name="inspection_scenarios.build_cross_sheet_formula",
        intentions=["Fórmulas preservadas sin evaluar"],
        features=["formulas", "cross_sheet"],
        expected_skeleton={
            "inspection": {
                "cells": [
                    {
                        "sheet": "Calc",
                        "coordinate": "B1",
                        "formula": "=Inputs!A2+Inputs!A3",
                    },
                    {"sheet": "Calc", "coordinate": "B2", "formula": "=SUM(Inputs!A2:A3)"},
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="named_range_formula",
        category="types",
        description="Fórmula que usa un rango nombrado",
        relative_workbook="workbooks/types/named_range_formula.xlsx",
        generator_name="inspection_scenarios.build_named_range_formula",
        intentions=["=SUM(Nums) sin evaluar"],
        features=["formulas", "defined_names"],
        expected_skeleton={
            "inspection": {
                "defined_names": [{"name": "Nums"}],
                "cells": [{"sheet": "Data", "coordinate": "C1", "formula": "=SUM(Nums)"}],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="xlsm_with_vba_stub",
        category="structural",
        description=(
            "XLSM con stub vbaProject.bin inocuo (no ejecutable); solo detección de presencia"
        ),
        relative_workbook="workbooks/structural/xlsm_with_vba_stub.xlsm",
        generator_name="inspection_scenarios.build_xlsm_with_vba_stub",
        intentions=[
            "Contiene xl/vbaProject.bin",
            "Nunca se ejecuta VBA ni se invoca Excel/COM",
            "has_vba_project=true",
        ],
        features=["xlsm", "vba_project_present", "no_macro_execution"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsm", "has_vba_project": True, "worksheet_count": 1},
                "warnings_contain": ["VBA_PROJECT_PRESENT"],
            }
        },
    ),
]


BUILDERS: dict[str, object] = {
    "inspection_scenarios.build_very_hidden_sheet": build_very_hidden_sheet,
    "inspection_scenarios.build_freeze_and_autofilter": build_freeze_and_autofilter,
    "inspection_scenarios.build_hidden_rows": build_hidden_rows,
    "inspection_scenarios.build_inflated_dimension": build_inflated_dimension,
    "inspection_scenarios.build_cell_physical_types": build_cell_physical_types,
    "inspection_scenarios.build_defined_names_variants": build_defined_names_variants,
    "inspection_scenarios.build_merged_variants": build_merged_variants,
    "inspection_scenarios.build_structured_tables_advanced": build_structured_tables_advanced,
    "inspection_scenarios.build_cross_sheet_formula": build_cross_sheet_formula,
    "inspection_scenarios.build_named_range_formula": build_named_range_formula,
    "inspection_scenarios.build_xlsm_with_vba_stub": build_xlsm_with_vba_stub,
}

SPECIAL_SAVERS: dict[str, object] = {
    "inspection_scenarios.build_xlsm_with_vba_stub": save_xlsm_with_vba_stub,
}
