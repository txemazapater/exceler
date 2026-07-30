from __future__ import annotations

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from tests.generators.workbook_factory import (
    ScenarioSpec,
    add_excel_table,
    bold_header,
    hide_column,
    new_workbook,
    write_matrix,
)


def build_empty_workbook(_seed: int) -> Workbook:
    return new_workbook(sheet_title="Empty")


def build_empty_sheet(_seed: int) -> Workbook:
    return new_workbook(sheet_title="Blank")


def build_simple_rectangular_table(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Data")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Id", "Name", "Amount"],
            [1, "Alpha", 10.5],
            [2, "Beta", 20.0],
            [3, "Gamma", 30.25],
        ],
    )
    bold_header(ws, 1, 3)
    return wb


def build_multi_sheet(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="One")
    ws1 = wb.active
    assert ws1 is not None
    write_matrix(ws1, [["A"], [1]])
    ws2 = wb.create_sheet("Two")
    write_matrix(ws2, [["B"], [2]])
    ws3 = wb.create_sheet("Three")
    write_matrix(ws3, [["C"], [3]])
    return wb


def build_leading_trailing_blank_rows(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Sparse")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            [None, None, None],
            [None, None, None],
            ["Code", "Label"],
            ["A1", "One"],
            ["A2", "Two"],
            [None, None],
            [None, None],
        ],
    )
    return wb


def build_interleaved_empty_columns(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Gaps")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["ColA", None, "ColC", None, "ColE"],
            ["x", None, "y", None, "z"],
            ["1", None, "2", None, "3"],
        ],
    )
    return wb


def build_title_above_header(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Titled")
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Informe de ventas — sintético"
    write_matrix(
        ws,
        [
            ["Producto", "Unidades", "Importe"],
            ["P-01", 3, 12.5],
            ["P-02", 1, 4.0],
        ],
        start_row=3,
    )
    bold_header(ws, 3, 3)
    return wb


def build_two_regions_one_sheet(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Regions")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["RegionA_Id", "RegionA_Name"],
            [1, "Norte"],
            [2, "Sur"],
        ],
        start_row=1,
        start_col=1,
    )
    write_matrix(
        ws,
        [
            ["RegionB_Code", "RegionB_Value"],
            ["X", 100],
            ["Y", 200],
        ],
        start_row=1,
        start_col=5,
    )
    return wb


def build_merged_cells(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Merged")
    ws = wb.active
    assert ws is not None
    ws.merge_cells("A1:C1")
    ws["A1"] = "Cabecera fusionada"
    write_matrix(
        ws,
        [
            ["Campo", "Valor"],
            ["A", 1],
            ["B", 2],
        ],
        start_row=3,
    )
    return wb


def build_hidden_sheet(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Visible")
    ws = wb.active
    assert ws is not None
    write_matrix(ws, [["VisibleCol"], ["ok"]])
    hidden = wb.create_sheet("HiddenData")
    write_matrix(hidden, [["SecretCol"], ["hidden-value"]])
    hidden.sheet_state = "hidden"
    return wb


def build_hidden_columns(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Cols")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Keep", "HideMe", "Keep2"],
            ["a", "secret", "b"],
        ],
    )
    hide_column(ws, 2)
    return wb


def build_formulas(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Calc")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Qty", "Price", "Total"],
            [2, 5, None],
            [3, 4, None],
        ],
    )
    ws["C2"] = "=A2*B2"
    ws["C3"] = "=A3*B3"
    ws["C4"] = "=SUM(C2:C3)"
    return wb


def build_cell_errors(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="Errors")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Label", "Value"],
            ["div0", "=1/0"],
            ["ref", "=Z99"],
            ["name", "=UNKNOWN()"],
        ],
    )
    return wb


def build_excel_structured_table(_seed: int) -> Workbook:
    wb = new_workbook(sheet_title="TableSheet")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Sku", "Qty"],
            ["S-1", 1],
            ["S-2", 2],
            ["S-3", 3],
        ],
    )
    bold_header(ws, 1, 2)
    add_excel_table(ws, name="InventoryTable", ref="A1:B4")
    return wb


def build_xlsm_container(_seed: int) -> Workbook:
    """Save as .xlsm extension without embedding vbaProject.bin.

    This fixture validates safe acceptance of the XLSM container type.
    It does not contain VBA macros and must never be executed via Excel/COM.
    """
    wb = new_workbook(sheet_title="MacroBook")
    ws = wb.active
    assert ws is not None
    write_matrix(
        ws,
        [
            ["Note", "Detail"],
            ["synthetic_xlsm", "extension_only_no_vba_project"],
        ],
    )
    wb.defined_names.add(DefinedName(name="FixtureFlag", attr_text="'MacroBook'!$A$2"))
    return wb


MINIMAL_SPECS: list[ScenarioSpec] = [
    ScenarioSpec(
        scenario_id="empty_workbook",
        category="minimal",
        description="Libro con una hoja vacía",
        relative_workbook="workbooks/minimal/empty_workbook.xlsx",
        generator_name="structural_scenarios.build_empty_workbook",
        intentions=["Una hoja sin celdas con valor"],
        features=["empty"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "has_vba_project": False, "worksheet_count": 1},
                "worksheets": [{"name": "Empty", "index": 0, "visibility": "visible"}],
            },
            "regions": {"reserved_for": "2B", "logical_tables": 0},
        },
    ),
    ScenarioSpec(
        scenario_id="empty_sheet",
        category="minimal",
        description="Hoja vacía con título explícito",
        relative_workbook="workbooks/minimal/empty_sheet.xlsx",
        generator_name="structural_scenarios.build_empty_sheet",
        intentions=["Hoja sin datos tabulares"],
        features=["empty_sheet"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "has_vba_project": False, "worksheet_count": 1},
                "worksheets": [{"name": "Blank", "index": 0, "visibility": "visible"}],
            },
            "regions": {"reserved_for": "2B", "logical_tables": 0},
        },
    ),
    ScenarioSpec(
        scenario_id="simple_rectangular_table",
        category="minimal",
        description="Tabla rectangular simple en A1",
        relative_workbook="workbooks/minimal/simple_rectangular_table.xlsx",
        generator_name="structural_scenarios.build_simple_rectangular_table",
        intentions=["Cabecera en fila 1", "Tres filas de datos"],
        features=["header_row", "rectangular"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "has_vba_project": False, "worksheet_count": 1},
                "worksheets": [
                    {
                        "name": "Data",
                        "index": 0,
                        "visibility": "visible",
                        "min_observed_cells": 12,
                    }
                ],
                "cells": [
                    {"sheet": "Data", "coordinate": "A1", "text": "Id"},
                    {"sheet": "Data", "coordinate": "A2", "value_kind": "integer"},
                ],
            },
            "regions": {"reserved_for": "2B", "logical_tables": 1},
        },
    ),
    ScenarioSpec(
        scenario_id="multi_sheet",
        category="structural",
        description="Tres hojas con una columna cada una",
        relative_workbook="workbooks/structural/multi_sheet.xlsx",
        generator_name="structural_scenarios.build_multi_sheet",
        intentions=["Exactamente tres hojas"],
        features=["multi_sheet"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "worksheet_count": 3},
                "worksheets": [
                    {"name": "One", "index": 0},
                    {"name": "Two", "index": 1},
                    {"name": "Three", "index": 2},
                ],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="leading_trailing_blank_rows",
        category="structural",
        description="Filas vacías antes y después de la tabla",
        relative_workbook="workbooks/structural/leading_trailing_blank_rows.xlsx",
        generator_name="structural_scenarios.build_leading_trailing_blank_rows",
        intentions=["La cabecera no está en la fila 1"],
        features=["offset_header", "blank_rows"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {"reserved_for": "2B", "logical_tables": 1},
        },
    ),
    ScenarioSpec(
        scenario_id="interleaved_empty_columns",
        category="structural",
        description="Columnas vacías intercaladas",
        relative_workbook="workbooks/structural/interleaved_empty_columns.xlsx",
        generator_name="structural_scenarios.build_interleaved_empty_columns",
        intentions=["Existen columnas sin cabecera entre columnas con datos"],
        features=["empty_columns"],
        expected_skeleton={"inspection": {"workbook": {"worksheet_count": 1}}},
    ),
    ScenarioSpec(
        scenario_id="title_above_header",
        category="structural",
        description="Título encima de la cabecera tabular",
        relative_workbook="workbooks/structural/title_above_header.xlsx",
        generator_name="structural_scenarios.build_title_above_header",
        intentions=["Fila 1 es título", "Cabecera en fila 3"],
        features=["title_row", "offset_header"],
        expected_skeleton={"inspection": {"workbook": {"worksheet_count": 1}}},
    ),
    ScenarioSpec(
        scenario_id="two_regions_one_sheet",
        category="structural",
        description="Dos regiones tabulares separadas en una hoja",
        relative_workbook="workbooks/structural/two_regions_one_sheet.xlsx",
        generator_name="structural_scenarios.build_two_regions_one_sheet",
        intentions=["Dos bloques independientes en la misma hoja"],
        features=["multi_region"],
        expected_skeleton={
            "inspection": {"workbook": {"worksheet_count": 1}},
            "regions": {"reserved_for": "2B", "logical_tables": 2},
        },
    ),
    ScenarioSpec(
        scenario_id="merged_cells",
        category="structural",
        description="Celdas combinadas sobre una tabla",
        relative_workbook="workbooks/structural/merged_cells.xlsx",
        generator_name="structural_scenarios.build_merged_cells",
        intentions=["Existe un rango combinado A1:C1"],
        features=["merged_cells"],
        expected_skeleton={
            "inspection": {
                "worksheets": [{"name": "Merged", "merged_ranges": ["A1:C1"]}],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="hidden_sheet",
        category="structural",
        description="Hoja oculta junto a una visible",
        relative_workbook="workbooks/structural/hidden_sheet.xlsx",
        generator_name="structural_scenarios.build_hidden_sheet",
        intentions=["Una hoja visible y una oculta"],
        features=["hidden_sheet"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsx", "has_vba_project": False, "worksheet_count": 2},
                "worksheets": [
                    {"name": "Visible", "index": 0, "visibility": "visible"},
                    {"name": "HiddenData", "index": 1, "visibility": "hidden"},
                ],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="hidden_columns",
        category="structural",
        description="Columna oculta entre columnas visibles",
        relative_workbook="workbooks/structural/hidden_columns.xlsx",
        generator_name="structural_scenarios.build_hidden_columns",
        intentions=["La columna B está oculta"],
        features=["hidden_columns"],
        expected_skeleton={
            "inspection": {
                "worksheets": [{"name": "Cols", "hidden_columns": ["B"]}],
            }
        },
    ),
    ScenarioSpec(
        scenario_id="formulas",
        category="types",
        description="Fórmulas de producto y suma",
        relative_workbook="workbooks/types/formulas.xlsx",
        generator_name="structural_scenarios.build_formulas",
        intentions=["C2/C3/C4 contienen fórmulas", "No se evalúan macros"],
        features=["formulas"],
        expected_skeleton={
            "inspection": {
                "cells": [
                    {"sheet": "Calc", "coordinate": "C2", "formula": "=A2*B2"},
                    {"sheet": "Calc", "coordinate": "C3", "formula": "=A3*B3"},
                    {"sheet": "Calc", "coordinate": "C4", "formula": "=SUM(C2:C3)"},
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="cell_errors",
        category="types",
        description="Celdas con fórmulas que producen errores",
        relative_workbook="workbooks/types/cell_errors.xlsx",
        generator_name="structural_scenarios.build_cell_errors",
        intentions=["Incluye #DIV/0!, #REF! y #NAME? potenciales"],
        features=["cell_errors"],
        expected_skeleton={
            "inspection": {
                "cells": [
                    {"sheet": "Errors", "coordinate": "B2", "formula": "=1/0"},
                    {"sheet": "Errors", "coordinate": "B3", "formula": "=Z99"},
                    {"sheet": "Errors", "coordinate": "B4", "formula": "=UNKNOWN()"},
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="excel_structured_table",
        category="structural",
        description="Tabla estructurada de Excel",
        relative_workbook="workbooks/structural/excel_structured_table.xlsx",
        generator_name="structural_scenarios.build_excel_structured_table",
        intentions=["Existe Table InventoryTable en A1:B4"],
        features=["excel_table"],
        expected_skeleton={
            "inspection": {
                "worksheets": [
                    {
                        "name": "TableSheet",
                        "tables": [
                            {
                                "name": "InventoryTable",
                                "ref": "A1:B4",
                                "totals_row_count": 0,
                            }
                        ],
                    }
                ]
            }
        },
    ),
    ScenarioSpec(
        scenario_id="xlsm_container",
        category="structural",
        description=(
            "Contenedor con extensión .xlsm sin proyecto VBA (sin vbaProject.bin); "
            "aceptación segura de la extensión, sin ejecución de macros"
        ),
        relative_workbook="workbooks/structural/xlsm_container.xlsm",
        generator_name="structural_scenarios.build_xlsm_container",
        intentions=[
            "Archivo guardado con extensión .xlsm",
            "No incluye vbaProject.bin",
            "No se ejecutan macros ni automatización COM en tests",
            "Incluye un nombre definido FixtureFlag para inspección estructural",
        ],
        features=["xlsm", "no_macro_execution", "no_vba_project", "defined_name"],
        expected_skeleton={
            "inspection": {
                "workbook": {"format": "xlsm", "has_vba_project": False, "worksheet_count": 1},
                "defined_names": [{"name": "FixtureFlag"}],
            }
        },
    ),
]


BUILDERS: dict[str, object] = {
    "structural_scenarios.build_empty_workbook": build_empty_workbook,
    "structural_scenarios.build_empty_sheet": build_empty_sheet,
    "structural_scenarios.build_simple_rectangular_table": build_simple_rectangular_table,
    "structural_scenarios.build_multi_sheet": build_multi_sheet,
    "structural_scenarios.build_leading_trailing_blank_rows": build_leading_trailing_blank_rows,
    "structural_scenarios.build_interleaved_empty_columns": build_interleaved_empty_columns,
    "structural_scenarios.build_title_above_header": build_title_above_header,
    "structural_scenarios.build_two_regions_one_sheet": build_two_regions_one_sheet,
    "structural_scenarios.build_merged_cells": build_merged_cells,
    "structural_scenarios.build_hidden_sheet": build_hidden_sheet,
    "structural_scenarios.build_hidden_columns": build_hidden_columns,
    "structural_scenarios.build_formulas": build_formulas,
    "structural_scenarios.build_cell_errors": build_cell_errors,
    "structural_scenarios.build_excel_structured_table": build_excel_structured_table,
    "structural_scenarios.build_xlsm_container": build_xlsm_container,
}
